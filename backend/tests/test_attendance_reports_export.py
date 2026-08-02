"""Attendance report queries + Excel/HTML builders (AR1–AR2).

The register matrix, the all-batches summary, and the file builders are tested
directly. The actual Chromium PDF render is not exercised here (no browser in
CI) — only the HTML the renderer consumes.
"""

import io
from datetime import date, datetime, timezone
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

from app.modules.attendance.models.attendance_models import DailyAttendance
from app.modules.attendance.services import attendance_export_service as ex
from app.modules.attendance.services import daily_service
from app.modules.lectures.models.lecture_models import Lecture
from app.modules.student.models.student_models import StudentBatchMapping

START, END = date(2026, 6, 1), date(2026, 6, 30)


def _lecture(seed_data, start):
    return Lecture(
        teacher_id=seed_data["teacher"].id, batch_id=seed_data["batch"].id,
        subject_id=seed_data["subject"].id, academic_year_id=seed_data["academic_year"].id,
        scheduled_start=start, scheduled_end=start.replace(hour=start.hour + 1),
        branch_id=seed_data["branch_a"].id, status="active", is_deleted=False,
    )


# ── queries ────────────────────────────────────────────────────────────────


@pytest.mark.usefixtures("seed_data")
async def test_batch_matrix_codes_and_totals(db_session, seed_data):
    db_session.add(StudentBatchMapping(
        student_id=seed_data["student"].id, batch_id=seed_data["batch"].id,
        branch_id=seed_data["branch_a"].id, status="active", is_deleted=False,
    ))
    # Two working days: 22nd and 23rd June (10:00 IST == 04:30 UTC).
    db_session.add(_lecture(seed_data, datetime(2026, 6, 22, 4, 30, tzinfo=timezone.utc)))
    db_session.add(_lecture(seed_data, datetime(2026, 6, 23, 4, 30, tzinfo=timezone.utc)))
    # Present on the 22nd only; 23rd has no row -> absent.
    db_session.add(DailyAttendance(
        student_id=seed_data["student"].id, branch_id=seed_data["branch_a"].id,
        attendance_date=date(2026, 6, 22), day_status="PRESENT", signoff="MISSING",
        source="BIOMETRIC",
    ))
    await db_session.commit()

    m = await daily_service.batch_matrix(
        session=db_session, branch_id=seed_data["branch_a"].id,
        batch_id=seed_data["batch"].id, start=START, end=END,
    )
    assert m["dates"] == [date(2026, 6, 22), date(2026, 6, 23)]
    assert m["student_count"] == 1
    row = m["students"][0]
    assert row["cells"] == ["P", "A"]
    assert row["present"] == 1
    assert row["attendance_pct"] == 50.0
    assert m["day_present"] == [1, 0]


@pytest.mark.usefixtures("seed_data")
async def test_batch_matrix_shows_punch_day_without_lecture(db_session, seed_data):
    """A real punch must never be hidden just because no lecture was scheduled
    that day (the moved-student / no-timetable batch case). The register column
    for the punch-day appears and the student reads present, not 0%."""
    db_session.add(StudentBatchMapping(
        student_id=seed_data["student"].id, batch_id=seed_data["batch"].id,
        branch_id=seed_data["branch_a"].id, status="active", is_deleted=False,
    ))
    # No lecture at all for the batch — but the student punched LATE on the 22nd.
    db_session.add(DailyAttendance(
        student_id=seed_data["student"].id, branch_id=seed_data["branch_a"].id,
        attendance_date=date(2026, 6, 22), day_status="LATE", signoff="COMPLETE",
        source="BIOMETRIC",
    ))
    await db_session.commit()

    m = await daily_service.batch_matrix(
        session=db_session, branch_id=seed_data["branch_a"].id,
        batch_id=seed_data["batch"].id, start=START, end=END,
    )
    assert m["dates"] == [date(2026, 6, 22)]  # punch-day is a column despite no lecture
    row = m["students"][0]
    assert row["cells"] == ["L"]
    assert row["present"] == 1
    assert row["attendance_pct"] == 100.0


@pytest.mark.usefixtures("seed_data")
async def test_daily_ledger_is_batch_independent(db_session, seed_data):
    """The ledger reads Layer-1 day facts only — a student with a record but no
    (current) batch mapping still appears, so history survives a batch move."""
    # No StudentBatchMapping at all for this student.
    db_session.add(DailyAttendance(
        student_id=seed_data["student"].id, branch_id=seed_data["branch_a"].id,
        attendance_date=date(2026, 6, 22), day_status="LATE",
        first_in=datetime(2026, 6, 22, 3, 59, tzinfo=timezone.utc),
        last_out=datetime(2026, 6, 22, 9, 0, tzinfo=timezone.utc),
        signoff="COMPLETE", source="BIOMETRIC",
    ))
    await db_session.commit()

    ledger = await daily_service.daily_ledger(
        session=db_session, branch_id=seed_data["branch_a"].id, start=START, end=END,
    )
    mine = [r for r in ledger if r["student_id"] == seed_data["student"].id]
    assert len(mine) == 1
    assert mine[0]["attendance_date"] == date(2026, 6, 22)
    assert mine[0]["day_status"] == "LATE"


@pytest.mark.usefixtures("seed_data")
async def test_branch_summary_aggregates(db_session, seed_data):
    db_session.add(StudentBatchMapping(
        student_id=seed_data["student"].id, batch_id=seed_data["batch"].id,
        branch_id=seed_data["branch_a"].id, status="active", is_deleted=False,
    ))
    db_session.add(_lecture(seed_data, datetime(2026, 6, 22, 4, 30, tzinfo=timezone.utc)))
    db_session.add(DailyAttendance(
        student_id=seed_data["student"].id, branch_id=seed_data["branch_a"].id,
        attendance_date=date(2026, 6, 22), day_status="PRESENT", signoff="NA",
        source="BIOMETRIC",
    ))
    await db_session.commit()

    rows = await daily_service.branch_summary(
        session=db_session, branch_id=seed_data["branch_a"].id, start=START, end=END,
    )
    seed_batch = next(r for r in rows if r["batch_id"] == seed_data["batch"].id)
    assert seed_batch["present"] == 1
    assert seed_batch["avg_pct"] == 100.0


# ── builders (pure) ─────────────────────────────────────────────────────────

def _matrix():
    return {
        "batch_id": "b1",
        "dates": [date(2026, 6, 22), date(2026, 6, 23)],
        "students": [
            {"student_id": "s1", "name": "Aarav Patil", "enrollment_number": "EN1",
             "cells": ["P", "A"], "present": 1, "working_days": 2, "attendance_pct": 50.0},
        ],
        "day_present": [1, 0],
        "student_count": 1,
    }


def _timeline():
    return [
        SimpleNamespace(
            attendance_date=date(2026, 6, 22),
            first_in=datetime(2026, 6, 22, 3, 59, tzinfo=timezone.utc),
            last_out=None, day_status="PRESENT", signoff="MISSING",
        ),
    ]


def _summary():
    return {"working_days": 2, "present_days": 1, "absent_days": 1, "attendance_pct": 50.0}


def test_student_xlsx_has_timeline(seed_data=None):
    data = ex.student_xlsx(
        brand="MSA", student_name="Aarav Patil", start=START, end=END,
        summary=_summary(), timeline=_timeline(), tz_name="Asia/Kolkata",
    )
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Attendance"]
    flat = [c.value for row in ws.iter_rows() for c in row if c.value is not None]
    assert "Aarav Patil" in str(flat)
    assert "2026-06-22" in str(flat)
    assert "09:29" in str(flat)  # 03:59 UTC -> 09:29 IST


def test_batch_xlsx_matrix_cells():
    data = ex.batch_xlsx(brand="MSA", batch_name="NEET-12-B", start=START, end=END, matrix=_matrix())
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Register"]
    flat = [c.value for row in ws.iter_rows() for c in row]
    assert "Aarav Patil" in flat
    assert "P" in flat and "A" in flat


def test_all_batches_xlsx_summary_plus_sheet():
    summaries = [{
        "batch_name": "NEET-12-B", "batch_code": "N12B", "student_count": 1,
        "working_days": 2, "present": 1, "total_slots": 2, "avg_pct": 50.0,
    }]
    data = ex.all_batches_xlsx(
        brand="MSA", start=START, end=END, summaries=summaries,
        matrices={"b1": _matrix()}, batch_names={"b1": "NEET-12-B"},
    )
    wb = load_workbook(io.BytesIO(data))
    assert "Summary" in wb.sheetnames
    assert "NEET-12-B" in wb.sheetnames  # per-batch detail sheet
    summary_flat = [c.value for row in wb["Summary"].iter_rows() for c in row]
    assert "NEET-12-B" in summary_flat


def _ledger():
    return [
        {"student_id": "s1", "name": "Aarav Patil", "enrollment_number": "EN1",
         "attendance_date": date(2026, 6, 22),
         "first_in": datetime(2026, 6, 22, 3, 59, tzinfo=timezone.utc),
         "last_out": datetime(2026, 6, 22, 9, 0, tzinfo=timezone.utc),
         "day_status": "LATE", "signoff": "COMPLETE", "source": "BIOMETRIC"},
    ]


def test_daily_ledger_xlsx_has_rows():
    data = ex.daily_ledger_xlsx(
        brand="MSA", start=START, end=END, ledger=_ledger(), tz_name="Asia/Kolkata",
    )
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Daily ledger"]
    flat = [c.value for row in ws.iter_rows() for c in row if c.value is not None]
    assert "Aarav Patil" in str(flat)
    assert "2026-06-22" in str(flat)
    assert "09:29" in str(flat)  # 03:59 UTC -> 09:29 IST


def test_daily_ledger_html_contains_data():
    s = ex.daily_ledger_html(
        brand="MSA", start=START, end=END, ledger=_ledger(), tz_name="Asia/Kolkata",
    )
    assert "daily attendance ledger" in s
    assert "Aarav Patil" in s and "LATE" in s


def test_html_builders_contain_data():
    s = ex.student_html(
        brand="MSA", student_name="Aarav Patil", start=START, end=END,
        summary=_summary(), timeline=_timeline(), tz_name="Asia/Kolkata",
    )
    assert "Aarav Patil" in s and "50.0%" in s

    b = ex.batch_html(brand="MSA", batch_name="NEET-12-B", start=START, end=END, matrix=_matrix())
    assert "NEET-12-B" in b and "Present / day" in b
