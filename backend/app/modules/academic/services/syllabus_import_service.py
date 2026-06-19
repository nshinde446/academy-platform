import io
import re
import uuid
from typing import Any

from fastapi import HTTPException, UploadFile, status
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.modules.academic import curriculum
from app.modules.academic.models.academic_models import (
    Chapter,
    Subject,
    Subtopic,
    Topic,
)
from app.modules.academic.repositories import academic_repository
from app.modules.audit.services import audit_service

# Header normalization: accept common variants.
HEADER_ALIASES = {
    "subject": "subject",
    "subject name": "subject",
    "chapter": "chapter",
    "chapter name": "chapter",
    "topic": "topic",
    "topic name": "topic",
    "subtopic": "subtopic",
    "subtopic name": "subtopic",
    "sub-topic": "subtopic",
    "sub topic": "subtopic",
}


def _normalize(header: str | None) -> str:
    if header is None:
        return ""
    return str(header).strip().lower()


def _parse_xlsx(content: bytes) -> list[dict[str, str]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb.active
    if ws is None:
        return []
    iterator = ws.iter_rows(values_only=True)
    try:
        raw_headers = list(next(iterator))
    except StopIteration:
        return []

    headers = [HEADER_ALIASES.get(_normalize(h), _normalize(h)) for h in raw_headers]

    out: list[dict[str, str]] = []
    for row in iterator:
        record: dict[str, str] = {}
        for i, value in enumerate(row):
            if i >= len(headers):
                break
            key = headers[i]
            if not key:
                continue
            record[key] = "" if value is None else str(value).strip()
        if record.get("subject"):
            out.append(record)
    return out


async def import_syllabus(
    session: AsyncSession,
    file: UploadFile,
    course_id: uuid.UUID,
    current_user_id: uuid.UUID,
    ip_address: str | None = None,
) -> dict[str, Any]:
    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only .xlsx files are supported for syllabus import.",
        )

    course = await academic_repository.get_course(session, course_id)
    if not course:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Course {course_id} not found.",
        )

    years = await academic_repository.list_academic_years(session, course.branch_id)
    if not years:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No academic year exists for this branch. Create one first.",
        )
    # Subjects/chapters/topics/subtopics carry academic_year_id; use the first
    # (oldest) year as the canonical owner. Syllabus is shared across years
    # for a course in practice.
    academic_year_id = years[0].id
    branch_id = course.branch_id

    content = await file.read()
    rows = _parse_xlsx(content)
    if not rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Sheet appears empty or missing a 'Subject' column.",
        )

    # Preload existing entities for this course so we can dedupe by (parent, name)
    # without N round-trips.
    existing_subjects = (
        await session.execute(
            select(Subject).where(
                Subject.course_id == course_id,
                Subject.branch_id == branch_id,
                Subject.is_deleted == False,
            )
        )
    ).scalars().all()
    subject_map: dict[str, uuid.UUID] = {s.name: s.id for s in existing_subjects}
    subject_ids = [s.id for s in existing_subjects]

    chapter_map: dict[tuple[uuid.UUID, str], uuid.UUID] = {}
    if subject_ids:
        existing_chapters = (
            await session.execute(
                select(Chapter).where(
                    Chapter.subject_id.in_(subject_ids),
                    Chapter.branch_id == branch_id,
                    Chapter.is_deleted == False,
                )
            )
        ).scalars().all()
        chapter_map = {(c.subject_id, c.name): c.id for c in existing_chapters}

    chapter_ids = list({cid for cid in chapter_map.values()})
    topic_map: dict[tuple[uuid.UUID, str], uuid.UUID] = {}
    if chapter_ids:
        existing_topics = (
            await session.execute(
                select(Topic).where(
                    Topic.chapter_id.in_(chapter_ids),
                    Topic.branch_id == branch_id,
                    Topic.is_deleted == False,
                )
            )
        ).scalars().all()
        topic_map = {(t.chapter_id, t.name): t.id for t in existing_topics}

    topic_ids = list({tid for tid in topic_map.values()})
    subtopic_map: dict[tuple[uuid.UUID, str], uuid.UUID] = {}
    if topic_ids:
        existing_subtopics = (
            await session.execute(
                select(Subtopic).where(
                    Subtopic.topic_id.in_(topic_ids),
                    Subtopic.branch_id == branch_id,
                    Subtopic.is_deleted == False,
                )
            )
        ).scalars().all()
        subtopic_map = {(s.topic_id, s.name): s.id for s in existing_subtopics}

    created = {
        "subjects_created": 0,
        "chapters_created": 0,
        "topics_created": 0,
        "subtopics_created": 0,
        "rows_processed": 0,
    }
    errors: list[str] = []

    def _code_for(name: str, prefix: str) -> str:
        # subjects/chapters/topics/subtopics columns include a `code` NOT NULL
        # for Subject (per schema). Derive a stable short code from the name
        # so we don't need to require it in the import file.
        cleaned = "".join(ch for ch in name.upper() if ch.isalnum())[:40]
        return cleaned or prefix

    for idx, row in enumerate(rows, start=2):  # row 1 is header
        subject_name = row.get("subject", "").strip()
        if not subject_name:
            continue

        try:
            # Subject
            if subject_name not in subject_map:
                subj = Subject(
                    branch_id=branch_id,
                    academic_year_id=academic_year_id,
                    course_id=course_id,
                    name=subject_name,
                    code=_code_for(subject_name, "SUBJ"),
                )
                session.add(subj)
                await session.flush()
                subject_map[subject_name] = subj.id
                created["subjects_created"] += 1
            subject_id = subject_map[subject_name]

            # Chapter (optional)
            chapter_name = row.get("chapter", "").strip()
            chapter_id: uuid.UUID | None = None
            if chapter_name:
                key = (subject_id, chapter_name)
                if key not in chapter_map:
                    ch = Chapter(
                        branch_id=branch_id,
                        academic_year_id=academic_year_id,
                        subject_id=subject_id,
                        name=chapter_name,
                        order=0,
                    )
                    session.add(ch)
                    await session.flush()
                    chapter_map[key] = ch.id
                    created["chapters_created"] += 1
                chapter_id = chapter_map[key]

            # Topic (optional)
            topic_name = row.get("topic", "").strip()
            topic_id: uuid.UUID | None = None
            if topic_name and chapter_id is not None:
                key = (chapter_id, topic_name)
                if key not in topic_map:
                    t = Topic(
                        branch_id=branch_id,
                        academic_year_id=academic_year_id,
                        chapter_id=chapter_id,
                        name=topic_name,
                        order=0,
                    )
                    session.add(t)
                    await session.flush()
                    topic_map[key] = t.id
                    created["topics_created"] += 1
                topic_id = topic_map[key]

            # Subtopic (optional)
            subtopic_name = row.get("subtopic", "").strip()
            if subtopic_name and topic_id is not None:
                key = (topic_id, subtopic_name)
                if key not in subtopic_map:
                    st = Subtopic(
                        branch_id=branch_id,
                        academic_year_id=academic_year_id,
                        topic_id=topic_id,
                        name=subtopic_name,
                        order=0,
                    )
                    session.add(st)
                    await session.flush()
                    subtopic_map[key] = st.id
                    created["subtopics_created"] += 1

            created["rows_processed"] += 1
        except Exception as exc:  # noqa: BLE001
            errors.append(f"Row {idx}: {exc}")

    await audit_service.log_action(
        session,
        user_id=current_user_id,
        action="IMPORT_SYLLABUS",
        table_name="academic",
        record_id=course_id,
        new_values={k: v for k, v in created.items()},
        ip_address=ip_address,
        branch_id=branch_id,
    )

    return {**created, "errors": errors}


# --- Curriculum backfill ---------------------------------------------------
# A student import only auto-populates a course's chapters/topics from the
# bundled master curriculum at the moment its subjects are first created. For
# courses whose subjects already existed (skeleton made before the curriculum
# feature shipped, or batches created manually), this backfills the curriculum
# onto subjects that still have none — additive only, keyed to the course's
# Target/syllabus, so it "maps the master syllabus to the students' Target".

def _candidate_syllabus_keys(course_code: str) -> list[str]:
    """Map a course code (as created by the student importer's TARGET_COURSE,
    possibly with a ``-NY`` duration suffix) to the curriculum syllabus key(s)
    to try. MHT-CET returns both streams; the per-subject sheet lookup picks the
    right one (Maths -> PCM, Biology -> PCB)."""
    c = re.sub(r"-\d+Y$", "", (course_code or "").strip().upper())
    if c == "NEET":
        return ["NEET"]
    if c == "JEE":
        return ["JEE"]
    if c in ("NEET-JEE", "BOTH", "PCMB"):
        return ["PCMB"]
    if c in ("MHT-CET", "MHTCET", "MHT-CET-PCM", "MHT-CET-PCB"):
        return ["MHT-CET-PCM", "MHT-CET-PCB"]
    return []


async def backfill_curriculum(
    session: AsyncSession,
    branch_id: uuid.UUID,
    current_user_id: uuid.UUID,
    ip_address: str | None = None,
) -> dict[str, int]:
    """Load the bundled master curriculum onto a branch's existing course
    subjects that have no chapters yet. Idempotent and additive — a subject that
    already has chapters (a curriculum a syllabus import loaded) is left
    untouched (§7.4). Returns counts of what was added."""
    courses = await academic_repository.list_courses(session, branch_id)
    subjects = await academic_repository.list_subjects(session, branch_id)
    by_course: dict[uuid.UUID, list] = {}
    for s in subjects:
        by_course.setdefault(s.course_id, []).append(s)

    courses_touched = 0
    subjects_filled = 0
    chapters_added = 0
    topics_added = 0

    for course in courses:
        keys = _candidate_syllabus_keys(course.code)
        if not keys:
            continue
        touched = False
        for subject in by_course.get(course.id, []):
            existing = await academic_repository.list_chapters(
                session, branch_id, subject.id
            )
            if existing:
                continue  # only fill empty subjects
            tree: tuple = ()
            for key in keys:
                tree = curriculum.chapters_for(key, subject.name)
                if tree:
                    break
            if not tree:
                continue
            for ch_order, (ch_name, topics) in enumerate(tree):
                chapter = await academic_repository.create_chapter(
                    session,
                    branch_id=branch_id,
                    academic_year_id=subject.academic_year_id,
                    subject_id=subject.id,
                    name=ch_name,
                    order=ch_order,
                )
                chapters_added += 1
                for t_order, t_name in enumerate(topics):
                    await academic_repository.create_topic(
                        session,
                        branch_id=branch_id,
                        academic_year_id=subject.academic_year_id,
                        chapter_id=chapter.id,
                        name=t_name,
                        order=t_order,
                    )
                    topics_added += 1
            subjects_filled += 1
            touched = True
        if touched:
            courses_touched += 1

    await audit_service.log_action(
        session,
        user_id=current_user_id,
        action="CURRICULUM_BACKFILL",
        table_name="chapters",
        record_id=uuid.uuid4(),
        new_values={
            "courses_touched": courses_touched,
            "subjects_filled": subjects_filled,
            "chapters_added": chapters_added,
            "topics_added": topics_added,
        },
        ip_address=ip_address,
        branch_id=branch_id,
    )
    return {
        "courses_touched": courses_touched,
        "subjects_filled": subjects_filled,
        "chapters_added": chapters_added,
        "topics_added": topics_added,
    }
