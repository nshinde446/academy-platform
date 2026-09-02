"""Rank-list exporters — branded PDF (notice-board / parents) and Excel.

Pure builders over the ``get_ranklist`` dict; reuses the attendance module's
headless-Chromium PDF renderer and the embedded MSA crest so the look matches
the day-attendance report.
"""

from __future__ import annotations

import html
import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.modules.attendance.services.attendance_export_service import (
    _logo_data_uri,
    render_html_to_pdf,
)

_HEAD_FILL = PatternFill("solid", fgColor="0F2947")
_HEAD_FONT = Font(bold=True, color="FFFFFF")
_BOLD = Font(bold=True)

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
PDF_MIME = "application/pdf"


def _esc(v: Any) -> str:
    return html.escape(str(v))


def _fmt_marks(m: float | None, total: float) -> str:
    if m is None:
        return "—"
    return f"{m:g} / {total:g}"


def ranklist_xlsx(*, brand: str, ranklist: dict) -> bytes:
    total = ranklist["total_marks"] or 0.0
    wb = Workbook()
    ws = wb.active
    ws.title = "Rank list"

    ws["A1"] = brand
    ws["A1"].font = Font(bold=True, size=14, color="003464")
    ws["A2"] = ranklist["test_name"]
    ws["A3"] = (
        f"Ranked {len(ranklist['ranked'])} · Absent {len(ranklist['absentees'])}"
        f" · Total marks {total:g}"
    )

    head = 5
    for c, h in enumerate(["Rank", "PRN", "Student Name", "Marks", "%"], start=1):
        cell = ws.cell(row=head, column=c, value=h)
        cell.font = _HEAD_FONT
        cell.fill = _HEAD_FILL

    r = head
    for row in ranklist["ranked"]:
        r += 1
        ws.cell(row=r, column=1, value=row["rank"])
        ws.cell(row=r, column=2, value=row.get("prn") or "")
        ws.cell(row=r, column=3, value=row["name"])
        ws.cell(row=r, column=4, value=_fmt_marks(row["marks_obtained"], total))
        ws.cell(row=r, column=5, value=(
            f"{row['percentage']:.1f}" if row["percentage"] is not None else ""
        ))
    for row in ranklist["absentees"]:
        r += 1
        ws.cell(row=r, column=1, value="—")
        ws.cell(row=r, column=2, value=row.get("prn") or "")
        ws.cell(row=r, column=3, value=row["name"])
        ws.cell(row=r, column=4, value="ABSENT")
        ws.cell(row=r, column=5, value="—")

    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


_CSS = """
@page { margin: 12mm 10mm; }
* { box-sizing: border-box; }
body { font-family: -apple-system, "Segoe UI", Arial, sans-serif; color:#111; font-size:10pt; margin:0; }
.brandbar { display:flex; align-items:center; gap:12px; border-bottom:2px solid #003464; padding-bottom:10px; }
.brandbar img { width:48px; height:48px; }
.brandbar h1 { font-size:18pt; margin:0; color:#003464; font-weight:800; }
.sub { color:#555; font-size:9.5pt; margin:8px 0 12px; }
table { border-collapse:collapse; width:100%; }
th, td { border:1px solid #d7dde4; padding:4pt 6pt; font-size:9.5pt; text-align:left; }
th { background:#0f2947; color:#fff; font-weight:700; }
tr:nth-child(even) td { background:#f7f9fb; }
td.c { text-align:center; }
.absent td { color:#c0392b; font-weight:700; }
.credit { text-align:right; color:#888; font-size:8pt; font-style:italic; margin-top:8px; }
"""


def ranklist_html(*, brand: str, ranklist: dict) -> str:
    total = ranklist["total_marks"] or 0.0
    logo = _logo_data_uri()
    logo_html = f"<img src='{logo}' alt=''>" if logo else ""
    body = [
        f"<div class='brandbar'>{logo_html}<h1>{_esc(brand)}</h1></div>",
        f"<p class='sub'><b>{_esc(ranklist['test_name'])}</b><br>"
        f"Ranked {len(ranklist['ranked'])} · Absent {len(ranklist['absentees'])}"
        f" · Total marks {total:g}</p>",
        "<table><tr><th>Rank</th><th>PRN</th><th>Student Name</th>"
        "<th>Marks</th><th>%</th></tr>",
    ]
    for row in ranklist["ranked"]:
        body.append(
            f"<tr><td class='c'>{row['rank']}</td>"
            f"<td>{_esc(row.get('prn') or '—')}</td>"
            f"<td>{_esc(row['name'])}</td>"
            f"<td class='c'>{_esc(_fmt_marks(row['marks_obtained'], total))}</td>"
            f"<td class='c'>{row['percentage']:.1f}</td></tr>"
            if row["percentage"] is not None else ""
        )
    for row in ranklist["absentees"]:
        body.append(
            f"<tr class='absent'><td class='c'>—</td>"
            f"<td>{_esc(row.get('prn') or '—')}</td>"
            f"<td>{_esc(row['name'])}</td>"
            f"<td class='c'>ABSENT</td><td class='c'>—</td></tr>"
        )
    body.append("</table>")
    body.append("<div class='credit'>Powered by EduPulse Technologies</div>")
    return (
        f"<!DOCTYPE html><html><head><meta charset='utf-8'>"
        f"<style>{_CSS}</style></head><body>{''.join(body)}</body></html>"
    )


async def ranklist_pdf(*, brand: str, ranklist: dict) -> bytes:
    return await render_html_to_pdf(ranklist_html(brand=brand, ranklist=ranklist))
