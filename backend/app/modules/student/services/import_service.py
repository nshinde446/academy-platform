import asyncio
import csv
import io
import re
import uuid
from collections import Counter
from datetime import date, datetime
from typing import Any

from fastapi import HTTPException, UploadFile, status
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.modules.academic import curriculum
from app.modules.academic.repositories import academic_repository
from app.modules.audit.services import audit_service
from app.modules.batch.models.batch_models import Batch
from app.modules.student.models.student_models import (
    Parent,
    Student,
    StudentBatchMapping,
    StudentImportJob,
)

# Reserved key the parsers stash the true source row number under. It is not a
# real column, so _row_to_student_kwargs (which only maps known headers) ignores
# it and it never reaches the Student model.
_ROW_KEY = "__row__"

# Excel/CSV column -> Student field
COLUMN_MAPPING = {
    "name": "name",
    "roll no": "enrollment_number",
    "roll_no": "enrollment_number",
    "rollno": "enrollment_number",
    "enrollment no": "enrollment_number",
    "enrollment_number": "enrollment_number",
    "email": "email",
    "phone": "phone",
    "parent mobile": "parent_mobile",
    "parent_mobile": "parent_mobile",
    "gender": "gender",
    "district": "district",
    "caste": "caste",
    "username": "username",
    # Date of birth — coerced to a real ``date`` in the build loop (multi-format).
    "dob": "date_of_birth",
    "date of birth": "date_of_birth",
    "date_of_birth": "date_of_birth",
    "birth date": "date_of_birth",
    "birthdate": "date_of_birth",
    # Enrollment / admission date — also coerced to a ``date`` in the build loop.
    "enrollment date": "enrollment_date",
    "enrollment_date": "enrollment_date",
    "admission date": "enrollment_date",
    "admission_date": "enrollment_date",
    "joining date": "enrollment_date",
    "date of joining": "enrollment_date",
    # Fees badge on the roster (paid|due|overdue|partial) — normalized in the loop.
    "fees": "fees_status",
    "fee status": "fees_status",
    "fees status": "fees_status",
    "fees_status": "fees_status",
    "payment status": "fees_status",
    # Original ambition kept when a 9th/10th aspirant is stored as Foundation (§3).
    "aspiration": "aspiration_target",
    "aspiration target": "aspiration_target",
    "aspiration_target": "aspiration_target",
    "rfid": "rfid_number",
    "rfidnumber": "rfid_number",
    "rfid_number": "rfid_number",
    "rfid number": "rfid_number",
    # Per-row enrolment fields. Each row picks its own class / exam /
    # batch — the dialog no longer carries them.
    "class": "standard",
    "standard": "standard",
    "target": "target_exam",
    "target exam": "target_exam",
    "target_exam": "target_exam",
    # Coaching stream (PCM/PCB/PCMB) the student opts into — a real per-student
    # field, distinct from the board-syllabus override column.
    "stream": "stream",
    "stream (pcm/pcb/pcmb)": "stream",
    "batch": "_batch_code",
    "batch code": "_batch_code",
    "batch_code": "_batch_code",
}

# When a referenced batch does not exist yet we can create it on import.
# The Target column drives which course the new batch instantiates and a
# sensible default exam date (month/day, anchored to the academic year's
# end year) so /insights has a pace anchor. Admins can edit both later.
TARGET_COURSE: dict[str, tuple[str, str]] = {
    "NEET": ("NEET", "NEET Preparation"),
    "JEE-Main": ("JEE", "JEE Preparation"),
    "JEE-Advanced": ("JEE", "JEE Preparation"),
    "MHT-CET": ("MHT-CET", "MHT-CET Preparation"),
    "Both": ("NEET-JEE", "NEET + JEE Preparation"),
    "Foundation": ("FND", "Foundation Programme"),
    "Other": ("GEN", "General Programme"),
}
_FALLBACK_COURSE = ("GEN", "General Programme")

TARGET_EXAM_MONTH_DAY: dict[str, tuple[int, int]] = {
    "NEET": (5, 4),
    "JEE-Main": (4, 6),
    "JEE-Advanced": (5, 18),
    "MHT-CET": (4, 24),
    "Both": (5, 4),
}

# Subject skeletons per syllabus (design §2/§5). DEFAULTS the coaching can edit
# — §6 leaves Biology-vs-Botany/Zoology and the MHT-CET stream split as open
# decisions, so we only auto-create for the unambiguous tracks and skip the
# rest (no subjects rather than a wrong guess).
SUBJECT_SETS: dict[str, list[str]] = {
    "NEET": ["Physics", "Chemistry", "Botany", "Zoology"],
    "JEE": ["Physics", "Chemistry", "Mathematics"],
    "PCMB": ["Physics", "Chemistry", "Mathematics", "Biology"],
    "MHT-CET-PCM": ["Physics", "Chemistry", "Mathematics"],
    "MHT-CET-PCB": ["Physics", "Chemistry", "Biology"],
    # MHT-CET with no explicit stream: carry the *union* on the course (P/C are
    # shared; Maths is PCM-only; Biology is PCB-only) and let each student's
    # `stream` select their subset at read time.
    "MHT-CET": ["Physics", "Chemistry", "Mathematics", "Biology"],
    "FOUNDATION": ["Science", "Mathematics", "Mental Ability"],
}

# Target -> default syllabus key when the row has no explicit Syllabus. MHT-CET
# now maps to its union set (the course carries all four subjects; per-student
# `stream` filters). Other has no academic subject set.
TARGET_SYLLABUS: dict[str, str] = {
    "NEET": "NEET",
    "JEE-Main": "JEE",
    "JEE-Advanced": "JEE",
    "Both": "PCMB",
    "MHT-CET": "MHT-CET",
    "Foundation": "FOUNDATION",
}

# Free-text Syllabus values normalized to a SUBJECT_SETS key.
_SYLLABUS_ALIASES: dict[str, str] = {
    "neet": "NEET",
    "pcb": "NEET",
    "jee": "JEE",
    "pcm": "JEE",
    "pcmb": "PCMB",
    "both": "PCMB",
    "mht-cet-pcm": "MHT-CET-PCM",
    "mhtcet-pcm": "MHT-CET-PCM",
    "mht-cet-pcb": "MHT-CET-PCB",
    "mhtcet-pcb": "MHT-CET-PCB",
    "foundation": "FOUNDATION",
}

_SUBJECT_CODES: dict[str, str] = {
    "Physics": "PHY",
    "Chemistry": "CHE",
    "Mathematics": "MAT",
    "Biology": "BIO",
    "Botany": "BOT",
    "Zoology": "ZOO",
    "Science": "SCI",
    "Mental Ability": "MA",
}

# Subjects that satisfy a target's exam requirement, for §3 Target×Syllabus
# consistency: NEET needs a biology subject, JEE needs maths.
_BIO_SUBJECTS = {"biology", "botany", "zoology"}
_MATHS_SUBJECTS = {"mathematics", "maths"}


def _syllabus_key(syllabus: str | None, target: str | None) -> str | None:
    """Resolve the subject-set key: explicit Syllabus wins, else the Target
    default. Returns None when the set is ambiguous/unknown (skip subjects)."""
    if syllabus:
        return _SYLLABUS_ALIASES.get(syllabus.strip().lower())
    return TARGET_SYLLABUS.get(target or "")


def _subjects_for(key: str | None) -> list[str]:
    return SUBJECT_SETS.get(key or "", [])


def _subject_code(name: str) -> str:
    return _SUBJECT_CODES.get(name, name[:3].upper())


def _split_name(value: str) -> tuple[str, str]:
    parts = value.strip().split(maxsplit=1)
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], parts[1]


# Date-of-birth accepts several common spellings; coerced to a real ``date`` in
# the build loop (a Date column needs a date object, not a string, on Postgres).
_DOB_FORMATS = (
    "%Y-%m-%d",
    "%d/%m/%Y",
    "%d-%m-%Y",
    "%d.%m.%Y",
    "%Y/%m/%d",
    "%d %b %Y",
    "%d %B %Y",
)
# Fees status enum surfaced as the roster Paid/Due badge (see Student.fees_status).
_VALID_FEES = {"paid", "due", "overdue", "partial"}


def _parse_date(value: Any) -> date | None:
    """Best-effort parse of a DOB cell to a ``date``. Handles ISO, the common
    Indian DD/MM/YYYY spellings, and the 'YYYY-MM-DD 00:00:00' that openpyxl
    renders for date cells. Returns None when nothing matches (the caller warns
    and leaves the field blank rather than failing the whole row)."""
    if value in (None, ""):
        return None
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if ":" in s:  # drop an Excel datetime's trailing time component
        s = s.split(" ", 1)[0]
    try:
        return date.fromisoformat(s)
    except ValueError:
        pass
    for fmt in _DOB_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _tidy_name(value: str) -> str:
    """Title-case names that arrive all-upper or all-lower (common in coaching
    exports); leave deliberately mixed-case input (e.g. 'McDonald') untouched."""
    return value.title() if value and (value.isupper() or value.islower()) else value


def _tidy_phone(value: str) -> str:
    """Keep digits and a leading '+', dropping spaces/dashes/parens so phone and
    parent_mobile store one consistent shape."""
    return re.sub(r"[^\d+]", "", value)


def _coerce_typed_fields(kwargs: dict[str, Any], rownum: int) -> list[str]:
    """Coerce the optional typed fields read off a row (date_of_birth → date,
    fees_status → enum) to their stored form. An unparseable value is dropped
    (not fatal) and reported as a warning so the rest of the row still imports."""
    warns: list[str] = []
    for field, label in (
        ("date_of_birth", "Date of Birth"),
        ("enrollment_date", "Enrollment Date"),
    ):
        raw = kwargs.get(field)
        if raw is not None and not isinstance(raw, date):
            parsed = _parse_date(raw)
            if parsed is None:
                warns.append(
                    f"Row {rownum}: unrecognized {label} '{raw}' — left blank"
                )
                kwargs.pop(field, None)
            else:
                kwargs[field] = parsed
    fees = kwargs.get("fees_status")
    if fees is not None:
        norm = str(fees).strip().lower()
        if norm in _VALID_FEES:
            kwargs["fees_status"] = norm
        else:
            warns.append(
                f"Row {rownum}: unrecognized Fees status '{fees}' "
                f"(use paid/due/overdue/partial) — left blank"
            )
            kwargs.pop("fees_status", None)
    return warns


def _normalize(header: str) -> str:
    return header.strip().lower()


def _norm_code(code: str) -> str:
    """Matching key for a batch code: trimmed, whitespace-collapsed,
    case-insensitive. Keeps lookups forgiving of stray spacing / casing
    without rewriting the code's own punctuation (so e.g. ``JEE--11-A``
    and ``JEE-Main-11-A`` stay distinct and surface separately)."""
    return " ".join(code.split()).strip().lower()


def _course_for_target(target: str | None) -> tuple[str, str]:
    return TARGET_COURSE.get(target or "", _FALLBACK_COURSE)


def _suggested_exam_date(target: str | None, end_year: int | None) -> date | None:
    md = TARGET_EXAM_MONTH_DAY.get(target or "")
    if md is None or end_year is None:
        return None
    return date(end_year, md[0], md[1])


def _dominant_target(counter: Counter) -> str | None:
    if not counter:
        return None
    return counter.most_common(1)[0][0]


# --- Optional per-row batch override columns (design §5) --------------------
# These let an admin who already knows their structure pin a created batch's
# course name, duration, and intake year instead of everything being a generic
# 1-year course derived from Target. All optional; explicit value wins, else we
# fall back to the Target-derived defaults. They are read straight off the raw
# row (NOT via COLUMN_MAPPING) so they never reach the Student model.
_OVERRIDE_HEADERS = {
    "course_opt": "course_opt",
    "course opt": "course_opt",
    "course": "course_opt",
    "course name": "course_opt",
    "duration": "duration",
    "academic_year": "academic_year",
    "academic year": "academic_year",
    "syllabus": "syllabus",
}


def _row_overrides(row: dict[str, Any]) -> dict[str, str]:
    out: dict[str, str] = {}
    for raw_key, raw_value in row.items():
        if raw_key is None:
            continue
        key = _OVERRIDE_HEADERS.get(_normalize(raw_key))
        if key is None or key in out:
            continue
        value = str(raw_value).strip() if raw_value not in (None, "") else ""
        if value:
            out[key] = value
    return out


# Optional parent/guardian columns → a Parent row (NOT via COLUMN_MAPPING, so
# they never reach the Student model). parent_mobile stays a Student column and
# is used as the Parent's phone when no explicit parent phone is given.
_PARENT_HEADERS = {
    "parent name": "name",
    "parent_name": "name",
    "father name": "name",
    "father's name": "name",
    "fathers name": "name",
    "mother name": "name",
    "guardian name": "name",
    "parent relation": "relation",
    "parent_relation": "relation",
    "relation": "relation",
    "parent occupation": "occupation",
    "parent_occupation": "occupation",
    "occupation": "occupation",
    "parent email": "email",
    "parent_email": "email",
    "parent phone": "phone",
    "parent_phone": "phone",
}


def _parent_fields(row: dict[str, Any]) -> dict[str, str]:
    """First-seen non-empty parent column values off the raw row."""
    out: dict[str, str] = {}
    for raw_key, raw_value in row.items():
        if raw_key is None:
            continue
        key = _PARENT_HEADERS.get(_normalize(str(raw_key)))
        if key is None or key in out:
            continue
        value = str(raw_value).strip() if raw_value not in (None, "") else ""
        if value:
            out[key] = value
    return out


def _build_parent(
    row: dict[str, Any],
    student_id: uuid.UUID,
    branch_id: uuid.UUID,
    parent_mobile: str | None,
    import_id: uuid.UUID | None,
) -> "Parent | None":
    """Create a Parent row when the upload carries a parent/guardian name. Phone
    falls back to the student's parent_mobile; relation defaults to Guardian."""
    fields = _parent_fields(row)
    name = fields.get("name")
    if not name:
        return None
    phone = fields.get("phone") or parent_mobile
    if phone:
        phone = _tidy_phone(str(phone)) or None
    return Parent(
        student_id=student_id,
        branch_id=branch_id,
        name=name,
        relation=fields.get("relation") or "Guardian",
        phone=phone,
        email=fields.get("email"),
        occupation=fields.get("occupation"),
        import_id=import_id,
    )


def _merge_overrides(into: dict[str, str], row_overrides: dict[str, str]) -> None:
    """First-seen non-empty value per field wins, so a batch referenced by many
    rows gets one stable set of overrides regardless of row order."""
    for k, v in row_overrides.items():
        into.setdefault(k, v)


def _parse_duration(value: str | None) -> int | None:
    """'2 Years' | '2yr' | '2' | '1 Year' -> int. Bounded to 1..6."""
    if not value:
        return None
    m = re.search(r"\d+", str(value))
    if not m:
        return None
    n = int(m.group())
    return n if 1 <= n <= 6 else None


def _parse_ay_span(value: str | None) -> tuple[int | None, int | None]:
    """'2026-2028' -> (2026, 2028); '2026-27' -> (2026, 2027); '2026' ->
    (2026, None). The end is the calendar end of the span, so span duration =
    end - start (a 2026-2028 span is a 2-year programme)."""
    if not value:
        return None, None
    nums = re.findall(r"\d{2,4}", str(value))
    if not nums:
        return None, None
    start = int(nums[0])
    if len(nums) == 1:
        return start, None
    end = int(nums[1])
    if end < 100:  # '2026-27' -> 2027
        end = (start // 100) * 100 + end
    return start, end


def _effective_duration_and_start(
    overrides: dict[str, str] | None,
) -> tuple[int, int | None]:
    """Resolve (duration_years, start_year) from the override columns. Explicit
    Duration wins; otherwise an Academic_year *span* implies the duration. Start
    year comes from Academic_year when given (else None = use the picked year)."""
    overrides = overrides or {}
    duration = _parse_duration(overrides.get("duration"))
    ay_start, ay_end = _parse_ay_span(overrides.get("academic_year"))
    if duration is None and ay_start is not None and ay_end is not None:
        duration = max(ay_end - ay_start, 1)
    return (duration or 1), ay_start


def _course_code_for(target: str | None, duration: int) -> str:
    """A course's identity includes its duration (design §1: 'NEET 2-Year' and
    'NEET 1-Year' are different courses), so a >1-year course gets a distinct,
    duration-suffixed code. 1-year keeps the bare code (backward compatible)."""
    base, _ = _course_for_target(target)
    return f"{base}-{duration}Y" if duration > 1 else base


def _course_name_for(
    target: str | None, duration: int, course_opt: str | None
) -> str:
    if course_opt:
        return course_opt
    _, base_name = _course_for_target(target)
    return f"{base_name} ({duration}-Year)" if duration > 1 else base_name


def _ay_by_start_year(years: list, start_year: int):
    for y in years:
        if y.start_year == start_year:
            return y
    return None


# --- Cross-field row validation (design §3: catch contradictions, don't ----
# silently coerce). Errors make a row non-importable; warnings are surfaced
# but allowed through.
_ONE_YEAR_CLASSES = {"12", "dropper"}
_EXAM_TARGETS = {"NEET", "JEE-Main", "JEE-Advanced", "Both"}


def _validate_row_consistency(
    standard: str | None,
    target: str | None,
    overrides: dict[str, str] | None,
) -> tuple[list[str], list[str]]:
    """Return (errors, warnings) for one row given its class, target, and
    override columns. Mirrors the §3 matrix, including the Target×Syllabus
    conflicts when an explicit Syllabus is given."""
    errors: list[str] = []
    warnings: list[str] = []
    overrides = overrides or {}
    cls = (standard or "").strip().lower()
    tgt = (target or "").strip()

    duration = _parse_duration(overrides.get("duration"))
    ay_start, ay_end = _parse_ay_span(overrides.get("academic_year"))
    span = (ay_end - ay_start) if (ay_start is not None and ay_end is not None) else None
    # Effective length: explicit Duration wins, else an Academic_year span.
    eff_duration = duration if duration is not None else span

    # E1/E2: 12th and Droppers are single-year programmes (one year before
    # the exam), so a multi-year length is a contradiction.
    if eff_duration is not None and eff_duration > 1 and cls in _ONE_YEAR_CLASSES:
        label = "Droppers" if cls == "dropper" else "Class 12"
        errors.append(
            f"{label} must be a 1-Year programme, not {eff_duration}-Year"
        )

    # E3: an explicit Academic_year span must agree with an explicit Duration.
    if duration is not None and span is not None and span != duration:
        errors.append(
            f"Academic_year span ({span}-Year) doesn't match Duration ({duration}-Year)"
        )

    # W1: an 11th explicitly put in a 1-Year programme finishes mid-program;
    # allowed (deliberate crash course) but worth flagging.
    if duration == 1 and cls == "11":
        warnings.append(
            "Class 11 in a 1-Year programme (crash course?) — confirm this is intended"
        )

    # W2: 9th/10th can't sit NEET/JEE; enrol as-is but flag (the Foundation
    # remap that would set aspiration_target isn't built yet).
    if cls in {"9", "10"} and tgt in _EXAM_TARGETS:
        warnings.append(
            f"Class {standard} targeting {tgt} — Foundation remap not yet "
            f"supported, enrolling as-is"
        )

    # Target × Syllabus conflicts (§3): only when an explicit Syllabus is given.
    syllabus_raw = overrides.get("syllabus")
    if syllabus_raw:
        key = _SYLLABUS_ALIASES.get(syllabus_raw.strip().lower())
        if key is None:
            warnings.append(
                f"Unrecognized Syllabus '{syllabus_raw}' — no subjects will be "
                f"auto-created for this batch"
            )
        else:
            subjects = {s.lower() for s in SUBJECT_SETS.get(key, [])}
            has_bio = bool(subjects & _BIO_SUBJECTS)
            has_maths = bool(subjects & _MATHS_SUBJECTS)
            if tgt == "NEET" and not has_bio:
                errors.append(
                    f"Target NEET needs a Biology syllabus, but Syllabus "
                    f"'{syllabus_raw}' has none"
                )
            if tgt in {"JEE-Main", "JEE-Advanced"} and not has_maths:
                errors.append(
                    f"Target {tgt} needs Maths, but Syllabus '{syllabus_raw}' "
                    f"has none"
                )
            if tgt == "Both" and not (has_bio and has_maths):
                errors.append(
                    f"Target Both needs PCMB (incl. Biology and Maths), but "
                    f"Syllabus '{syllabus_raw}' is incomplete"
                )
            if tgt == "NEET" and has_maths:
                warnings.append(
                    "Target NEET with a Maths syllabus — extra Maths the "
                    "student won't sit (allowed but inefficient)"
                )

    return errors, warnings


# --- Batch-code override conflict (design §3 "batch identity collision"): -----
# when many rows share a batch code but disagree on the course/length/year the
# new batch should have, that's an error, not a silent first-seen-wins.
_OVERRIDE_LABELS = {
    "course_opt": "Course_opt",
    "duration": "Duration",
    "academic_year": "Academic_year",
}


def _canon_override(field: str, value: str) -> str | None:
    """Canonical form so cosmetic differences ('2 Years' vs '2') aren't read as
    a conflict; returns None when the value doesn't parse to anything."""
    if field == "duration":
        d = _parse_duration(value)
        return str(d) if d is not None else None
    if field == "academic_year":
        start, end = _parse_ay_span(value)
        return None if start is None else f"{start}-{end}"
    return value.strip().lower()


def _track_override_values(
    store: dict[str, set[str]], row_overrides: dict[str, str]
) -> None:
    for field, value in row_overrides.items():
        canon = _canon_override(field, value)
        if canon:
            store.setdefault(field, set()).add(canon)


def _override_conflict(field_values: dict[str, set[str]]) -> str | None:
    conflicted = [f for f, vals in field_values.items() if len(vals) > 1]
    if not conflicted:
        return None
    return "rows disagree on " + ", ".join(
        _OVERRIDE_LABELS.get(f, f) for f in conflicted
    )


def _pick_academic_year(years: list):
    """Deterministically choose the academic year new students enrol into.

    ``list_academic_years`` has no ORDER BY, so indexing ``years[0]`` picked an
    arbitrary row — with more than one year on a branch that silently pinned
    every imported student (and every derived batch's start year) to the wrong
    year, and could even differ between preview and import. Default to the most
    recent year by ``start_year``."""
    if not years:
        return None
    return max(years, key=lambda y: y.start_year)


def _parse_csv(content: bytes) -> list[dict[str, Any]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    out: list[dict[str, Any]] = []
    for row in reader:
        record = dict(row)
        # True source line (DictReader skips blank lines, so a running counter
        # would drift); used for accurate "Row N" diagnostics.
        record[_ROW_KEY] = reader.line_num
        out.append(record)
    return out


def _parse_xlsx(content: bytes) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb.active
    if ws is None:
        return []
    rows = ws.iter_rows(values_only=True)
    try:
        headers = [str(h) if h is not None else "" for h in next(rows)]
    except StopIteration:
        return []
    out: list[dict[str, Any]] = []
    for excel_row, row in enumerate(rows, start=2):  # header was row 1
        record = {
            headers[i]: ("" if row[i] is None else str(row[i]))
            for i in range(min(len(headers), len(row)))
        }
        if any(v for v in record.values()):
            # Real worksheet row number, so "Row N" matches what the admin
            # sees even when blank rows were skipped.
            record[_ROW_KEY] = excel_row
            out.append(record)
    return out


def _parse_file(filename: str | None, content: bytes) -> list[dict[str, Any]]:
    name = (filename or "").lower()
    if name.endswith(".csv"):
        return _parse_csv(content)
    if name.endswith((".xlsx", ".xls")):
        return _parse_xlsx(content)
    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail="Unsupported file type. Use .csv or .xlsx",
    )


def _row_to_student_kwargs(
    row: dict[str, Any], branch_id: uuid.UUID, academic_year_id: uuid.UUID
) -> dict[str, Any] | None:
    mapped: dict[str, Any] = {}
    for raw_key, raw_value in row.items():
        if raw_key is None:
            continue
        field = COLUMN_MAPPING.get(_normalize(raw_key))
        if field is None:
            continue
        value = (raw_value or "").strip() if isinstance(raw_value, str) else raw_value
        if value in ("", None):
            continue
        mapped[field] = value

    name_value = mapped.pop("name", None)
    if name_value:
        first, last = _split_name(str(name_value))
        mapped["first_name"] = _tidy_name(first)
        mapped["last_name"] = _tidy_name(last)

    if not mapped.get("first_name"):
        return None

    # Normalize phone shapes (digits + optional leading '+'); drop if nothing
    # usable is left (e.g. a stray 'N/A').
    for pkey in ("phone", "parent_mobile"):
        raw = mapped.get(pkey)
        if isinstance(raw, str):
            cleaned = _tidy_phone(raw)
            if cleaned:
                mapped[pkey] = cleaned
            else:
                mapped.pop(pkey, None)

    mapped["branch_id"] = branch_id
    mapped["academic_year_id"] = academic_year_id
    mapped.setdefault("last_name", "")
    return mapped


# Canonical import fields the column-mapping UI (T3) can target. ``key`` is a
# header string the parser already understands (it normalizes into
# COLUMN_MAPPING / _OVERRIDE_HEADERS / _PARENT_HEADERS); ``label`` is what the
# admin sees in the mapping dropdown.
IMPORT_FIELDS: list[dict[str, str]] = [
    {"key": "name", "label": "Name", "required": "1"},
    {"key": "class", "label": "Class"},
    {"key": "target", "label": "Target exam"},
    {"key": "batch", "label": "Batch"},
    {"key": "roll no", "label": "Roll No / Enrollment"},
    {"key": "email", "label": "Email"},
    {"key": "phone", "label": "Phone"},
    {"key": "parent mobile", "label": "Parent mobile"},
    {"key": "gender", "label": "Gender"},
    {"key": "district", "label": "District"},
    {"key": "caste", "label": "Caste"},
    {"key": "username", "label": "Username"},
    {"key": "rfid number", "label": "RFID number"},
    {"key": "dob", "label": "Date of birth"},
    {"key": "fees", "label": "Fees status"},
    {"key": "enrollment date", "label": "Enrollment date"},
    {"key": "stream", "label": "Stream"},
    {"key": "aspiration", "label": "Aspiration target"},
    {"key": "parent name", "label": "Parent name"},
    {"key": "relation", "label": "Parent relation"},
    {"key": "occupation", "label": "Parent occupation"},
    {"key": "parent phone", "label": "Parent phone"},
    {"key": "course_opt", "label": "Course (override)"},
    {"key": "duration", "label": "Duration (override)"},
    {"key": "academic_year", "label": "Academic year (override)"},
    {"key": "syllabus", "label": "Syllabus (override)"},
]


def _recognized_norm_keys() -> set[str]:
    return set(COLUMN_MAPPING) | set(_OVERRIDE_HEADERS) | set(_PARENT_HEADERS)


def _build_field_key_for_header() -> dict[str, str]:
    """Map every recognized normalized header → its canonical IMPORT_FIELDS key,
    so the grid collapses aliases (e.g. 'standard' and 'class' both → 'class';
    'date of birth' and 'dob' both → 'dob') onto one column per field."""
    key_by_student: dict[str, str] = {}
    key_by_override: dict[str, str] = {}
    key_by_parent: dict[str, str] = {}
    for f in IMPORT_FIELDS:
        k = f["key"]
        if k in COLUMN_MAPPING:
            key_by_student.setdefault(COLUMN_MAPPING[k], k)
        elif k in _OVERRIDE_HEADERS:
            key_by_override.setdefault(_OVERRIDE_HEADERS[k], k)
        elif k in _PARENT_HEADERS:
            key_by_parent.setdefault(_PARENT_HEADERS[k], k)
    out: dict[str, str] = {}
    for header, target in COLUMN_MAPPING.items():
        if target in key_by_student:
            out[header] = key_by_student[target]
    for header, target in _OVERRIDE_HEADERS.items():
        if target in key_by_override:
            out[header] = key_by_override[target]
    for header, target in _PARENT_HEADERS.items():
        if target in key_by_parent:
            out[header] = key_by_parent[target]
    return out


_FIELD_KEY_FOR_HEADER = _build_field_key_for_header()


def parse_import_rows(
    filename: str | None,
    content: bytes,
    column_map: dict[str, str] | None,
) -> tuple[list[str], list[dict[str, Any]]]:
    """Parse an upload into the editable grid's shape: the canonical fields
    present (in IMPORT_FIELDS order) and one ``{index, row_number, values}`` per
    data row, where ``values`` is keyed by canonical field key. Column-map (or
    the file's own canonical headers) decides which file column feeds each field."""
    raw = _apply_column_map(_parse_file(filename, content), column_map)
    present: set[str] = set()
    rows: list[dict[str, Any]] = []
    for idx, r in enumerate(raw):
        rownum = _row_number(r, idx + 2)
        values: dict[str, str] = {}
        for k, v in r.items():
            if k is None or k == _ROW_KEY:
                continue
            field_key = _FIELD_KEY_FOR_HEADER.get(_normalize(str(k)))
            if field_key is None or field_key in values:
                continue
            val = str(v).strip() if v not in (None, "") else ""
            if val:
                values[field_key] = val
                present.add(field_key)
        rows.append({"index": idx, "row_number": rownum, "values": values})
    fields = [f["key"] for f in IMPORT_FIELDS if f["key"] in present]
    return fields, rows


def _file_headers(rows: list[dict[str, Any]]) -> list[str]:
    """The upload's header columns in first-seen order (sans the row-number key)."""
    out: list[str] = []
    for row in rows:
        for k in row.keys():
            if k is None or k == _ROW_KEY:
                continue
            s = str(k)
            if s not in out:
                out.append(s)
        if out:
            break  # the first data row carries every header
    return out


def _suggest_mapping(headers: list[str]) -> dict[str, str | None]:
    """Best-guess file-header → canonical-field-key map: a header that already
    normalizes to a known field maps to it; anything else is left unmapped."""
    recognized = _recognized_norm_keys()
    return {
        h: (_normalize(h) if _normalize(h) in recognized else None) for h in headers
    }


def _apply_column_map(
    rows: list[dict[str, Any]], column_map: dict[str, str] | None
) -> list[dict[str, Any]]:
    """Rename each row's headers per an explicit file-header → canonical-key map
    (from the mapping UI). Headers not in the map keep their original name, so
    the existing fixed-alias detection still applies to them. Empty/None mapping
    targets drop the column."""
    if not column_map:
        return rows
    out: list[dict[str, Any]] = []
    for row in rows:
        new: dict[str, Any] = {}
        for k, v in row.items():
            if k == _ROW_KEY:
                new[k] = v
                continue
            if k is not None and str(k) in column_map:
                target = column_map[str(k)]
                if not target:
                    continue  # explicitly skipped column
                new[target] = v
            else:
                new[k] = v
        out.append(new)
    return out


async def detect_columns(
    file: UploadFile,
) -> dict[str, Any]:
    """For the mapping step: the upload's headers, a suggested mapping, and the
    catalog of fields an admin can map them to."""
    content = await file.read()
    rows = _parse_file(file.filename, content)
    headers = _file_headers(rows)
    return {
        "headers": headers,
        "suggested": _suggest_mapping(headers),
        "fields": IMPORT_FIELDS,
    }


def _unrecognized_columns(rows: list[dict[str, Any]]) -> list[str]:
    """File header columns that map to no known field — their data would be
    silently dropped on import. Returned in original spelling, de-duped in
    first-seen order. The reserved row-number key and blank headers are ignored."""
    recognized = set(COLUMN_MAPPING) | set(_OVERRIDE_HEADERS) | set(_PARENT_HEADERS)
    seen: set[str] = set()
    out: list[str] = []
    for row in rows:
        for raw_key in row.keys():
            if raw_key is None or raw_key == _ROW_KEY:
                continue
            norm = _normalize(str(raw_key))
            if not norm or norm in recognized or norm in seen:
                continue
            seen.add(norm)
            out.append(str(raw_key).strip())
    return out


async def _load_batch_index(
    session: AsyncSession, branch_id: uuid.UUID
) -> dict[str, uuid.UUID]:
    """Map normalized batch code -> batch id for the branch's live batches."""
    result = await session.execute(
        select(Batch).where(
            Batch.branch_id == branch_id,
            Batch.is_deleted == False,  # noqa: E712
        )
    )
    return {_norm_code(b.code): b.id for b in result.scalars().all()}


async def _get_or_create_course(
    session: AsyncSession,
    branch_id: uuid.UUID,
    code: str,
    name: str,
    duration_years: int = 1,
    import_id: uuid.UUID | None = None,
):
    """Course identity is code (which encodes duration, design §1). Reuse an
    existing course with this code; otherwise create it and tag the new row with
    ``import_id`` so undo can reclaim a course this import brought into being."""
    courses = await academic_repository.list_courses(session, branch_id)
    for c in courses:
        if c.code.strip().lower() == code.lower():
            return c
    course = await academic_repository.create_course(
        session,
        branch_id=branch_id,
        name=name,
        code=code,
        duration_years=duration_years,
    )
    if import_id is not None:
        course.import_id = import_id
    return course


async def _ensure_subject_skeleton(
    session: AsyncSession,
    branch_id: uuid.UUID,
    course,
    academic_year,
    syllabus_key: str | None,
    import_id: uuid.UUID | None,
) -> int:
    """Create the course's subject skeleton from the resolved syllabus (design
    §4 step 3). The §8 protection: only ever create when the course has *no*
    subjects yet — never overwrite an existing skeleton or a curriculum that a
    syllabus import has since loaded (§7.4). Returns how many were created.

    Matched by ``(course, name)`` ignoring AY, so the later syllabus import
    finds and reuses these subjects when it attaches chapters."""
    names = _subjects_for(syllabus_key)
    if not names:
        return 0
    existing = await academic_repository.list_subjects(session, branch_id, course.id)
    if existing:
        return 0

    from app.modules.academic.models.academic_models import (
        Chapter,
        Subject,
        Topic,
    )

    # Build the whole subject -> chapter -> topic tree in memory (ids
    # pre-generated so children can reference parents) and insert it level by
    # level — subjects, then chapters, then topics — flushing between levels so
    # every parent row exists before its children. (One bulk add_all of the
    # whole tree doesn't guarantee insert order without ORM relationships, which
    # Postgres rejects as a FK violation; SQLite silently allowed it.) Still far
    # faster than the old per-row create+flush. Tagged with import_id so undo can
    # reclaim them. Foundation/Other have no curriculum and add nothing.
    subjects: list = []
    chapters: list = []
    topics_rows: list = []
    for name in names:
        subject_id = uuid.uuid4()
        subjects.append(
            Subject(
                id=subject_id,
                branch_id=branch_id,
                academic_year_id=academic_year.id,
                course_id=course.id,
                name=name,
                code=_subject_code(name),
                import_id=import_id,
            )
        )
        for ch_order, (ch_name, topics) in enumerate(
            curriculum.chapters_for(syllabus_key or "", name)
        ):
            chapter_id = uuid.uuid4()
            chapters.append(
                Chapter(
                    id=chapter_id,
                    branch_id=branch_id,
                    academic_year_id=academic_year.id,
                    subject_id=subject_id,
                    name=ch_name,
                    order=ch_order,
                    import_id=import_id,
                )
            )
            for t_order, t_name in enumerate(topics):
                topics_rows.append(
                    Topic(
                        id=uuid.uuid4(),
                        branch_id=branch_id,
                        academic_year_id=academic_year.id,
                        chapter_id=chapter_id,
                        name=t_name,
                        order=t_order,
                        import_id=import_id,
                    )
                )
    session.add_all(subjects)
    await session.flush()
    session.add_all(chapters)
    await session.flush()
    session.add_all(topics_rows)
    await session.flush()
    return len(names)


async def _create_derived_batch(
    session: AsyncSession,
    branch_id: uuid.UUID,
    code: str,
    target: str | None,
    academic_year,
    current_user_id: uuid.UUID,
    ip_address: str | None,
    import_id: uuid.UUID | None = None,
    overrides: dict[str, str] | None = None,
    years: list | None = None,
) -> tuple[Any, int]:
    """Create a batch for an unknown code and its course's subject skeleton.
    Course, duration, intake year and exam date come from the optional override
    columns (Course_opt / Duration / Academic_year / Syllabus) when present,
    else are derived from the dominant Target. Batch (and any subjects it
    creates) are tagged with ``import_id`` so an "undo import" can reclaim them.
    Returns ``(batch, subjects_created)``."""
    from app.modules.batch.schemas.batch_schemas import BatchCreate
    from app.modules.batch.services import batch_service

    duration, start_year = _effective_duration_and_start(overrides)

    start_ay = academic_year
    if start_year is not None:
        start_ay = _ay_by_start_year(years or [], start_year)
        if start_ay is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    f"needs an academic year starting at {start_year} "
                    f"(create it first)"
                ),
            )

    course_code = _course_code_for(target, duration)
    course_name = _course_name_for(
        target, duration, (overrides or {}).get("course_opt")
    )
    course = await _get_or_create_course(
        session,
        branch_id,
        course_code,
        course_name,
        duration_years=duration,
        import_id=import_id,
    )
    data = BatchCreate(
        branch_id=branch_id,
        start_academic_year_id=start_ay.id,
        course_id=course.id,
        name=code,
        code=code,
        capacity=30,
        # Exam happens in the calendar year the programme ends.
        target_exam_date=_suggested_exam_date(
            target, start_ay.start_year + duration
        ),
    )
    batch = await batch_service.create_batch(
        session, data, current_user_id, ip_address
    )
    if import_id is not None:
        batch.import_id = import_id
    syllabus_key = _syllabus_key((overrides or {}).get("syllabus"), target)
    subjects_created = await _ensure_subject_skeleton(
        session, branch_id, course, start_ay, syllabus_key, import_id
    )
    await session.flush()
    return batch, subjects_created


def _missing_batch_blocker(
    target: str | None,
    default_start: int,
    courses_by_code: dict[str, Any],
    ay_start_years: set[int],
    overrides: dict[str, str] | None = None,
) -> str | None:
    """Dry-run the academic-year preconditions ``create_batch`` enforces: the
    batch needs both its start year and its end year (``start + duration - 1``)
    to be available. ``ay_start_years`` includes years the import will itself
    auto-create, so this only fires for years that genuinely won't exist.

    Honors the override columns (Duration / Academic_year) and the rule that an
    existing course with the resolved code pins the duration (design §1)."""
    duration, start_year_override = _effective_duration_and_start(overrides)
    start_year = (
        start_year_override
        if start_year_override is not None
        else default_start
    )
    if start_year not in ay_start_years:
        return f"needs an academic year starting at {start_year} (create it first)"
    course_code = _course_code_for(target, duration)
    existing = courses_by_code.get(course_code.lower())
    eff_duration = (
        max(int(existing.duration_years or 1), 1) if existing else duration
    )
    end_year_start = start_year + eff_duration - 1
    if end_year_start not in ay_start_years:
        return f"needs an academic year starting at {end_year_start} (create it first)"
    return None


def _row_number(row: dict[str, Any], fallback: int) -> int:
    n = row.get(_ROW_KEY)
    return n if isinstance(n, int) else fallback


def _dup_keys(kwargs: dict[str, Any]) -> tuple[str | None, str | None]:
    """Natural keys used to detect a student already on file: enrolment number
    and email, both case-/space-insensitive. Either may be absent."""
    enr = kwargs.get("enrollment_number")
    em = kwargs.get("email")
    return (
        str(enr).strip().lower() if enr else None,
        str(em).strip().lower() if em else None,
    )


async def _load_student_dedup_keys(
    session: AsyncSession, branch_id: uuid.UUID
) -> tuple[set[str], set[str]]:
    """Existing enrolment numbers and emails for the branch's live students.

    Neither column is unique at the DB level, so without this an admin who
    re-uploads the same file just doubles every student (the real incident that
    produced 2000 rows from a 1000-row file). Seeded into the per-import 'seen'
    sets so re-imports and within-file repeats are skipped, not duplicated."""
    result = await session.execute(
        select(Student.enrollment_number, Student.email).where(
            Student.branch_id == branch_id,
            Student.is_deleted == False,  # noqa: E712
        )
    )
    enrols: set[str] = set()
    emails: set[str] = set()
    for enr, em in result.all():
        if enr:
            enrols.add(enr.strip().lower())
        if em:
            emails.add(em.strip().lower())
    return enrols, emails


def _fuzzy_keys(kwargs: dict[str, Any]) -> set[str]:
    """Soft identity keys for catching the same human re-entered *without* a roll
    number: (name + phone) and (name + DOB). Used only for a WARNING — never a
    hard skip, since genuine homonyms exist. Phone/DOB are normalized so messy
    spellings still match."""
    first = str(kwargs.get("first_name") or "").strip().lower()
    last = str(kwargs.get("last_name") or "").strip().lower()
    name = f"{first} {last}".strip()
    if not name:
        return set()
    keys: set[str] = set()
    phone = kwargs.get("phone")
    if phone:
        digits = _tidy_phone(str(phone))
        if digits:
            keys.add(f"{name}|p|{digits}")
    dob = _parse_date(kwargs.get("date_of_birth"))
    if dob is not None:
        keys.add(f"{name}|d|{dob.isoformat()}")
    return keys


async def _load_fuzzy_keys(
    session: AsyncSession, branch_id: uuid.UUID
) -> set[str]:
    """Fuzzy identity keys (name+phone / name+DOB) for the branch's live
    students — seeded into the per-import 'seen' set so an existing student
    re-entered without a roll number is flagged as a possible duplicate."""
    result = await session.execute(
        select(
            Student.first_name,
            Student.last_name,
            Student.phone,
            Student.date_of_birth,
        ).where(
            Student.branch_id == branch_id,
            Student.is_deleted == False,  # noqa: E712
        )
    )
    keys: set[str] = set()
    for first, last, phone, dob in result.all():
        keys |= _fuzzy_keys(
            {
                "first_name": first,
                "last_name": last,
                "phone": phone,
                "date_of_birth": dob,
            }
        )
    return keys


_BULK_CHUNK = 500


async def _bulk_insert_students(
    session: AsyncSession,
    pending: list[tuple[int, Student, "StudentBatchMapping | None", "Parent | None"]],
    on_progress=None,
) -> tuple[int, list[tuple[int, str]]]:
    """Insert validated students (+ their batch mappings and parent rows) in
    chunks — one flush per chunk instead of per row, which is what kept large
    imports under the request timeout. If a chunk's flush fails (a stray
    IntegrityError), that chunk is retried row-by-row in savepoints so a single
    bad row is isolated and reported, not fatal. ``on_progress`` (if given) is
    awaited after each chunk with the running processed count, for a job progress
    bar. Returns (imported_count, [(rownum, error), ...])."""
    imported = 0
    processed = 0
    skips: list[tuple[int, str]] = []
    for start in range(0, len(pending), _BULK_CHUNK):
        chunk = pending[start : start + _BULK_CHUNK]
        students = [student for _, student, _, _ in chunk]
        mappings = [m for _, _, m, _ in chunk if m is not None]
        parents = [p for _, _, _, p in chunk if p is not None]
        try:
            async with session.begin_nested():
                # Students first, then their children (mappings + parents both
                # FK the student) — parent rows must land before children
                # (Postgres enforces this; SQLite doesn't).
                session.add_all(students)
                await session.flush()
                if mappings:
                    session.add_all(mappings)
                if parents:
                    session.add_all(parents)
                if mappings or parents:
                    await session.flush()
            imported += len(chunk)
        except Exception:  # noqa: BLE001 - isolate the offending row(s)
            for rownum, student, mapping, parent in chunk:
                try:
                    async with session.begin_nested():
                        session.add(student)
                        await session.flush()
                        if mapping is not None:
                            session.add(mapping)
                        if parent is not None:
                            session.add(parent)
                        if mapping is not None or parent is not None:
                            await session.flush()
                    imported += 1
                except Exception as exc:  # noqa: BLE001
                    skips.append((rownum, str(exc)))
        processed += len(chunk)
        if on_progress is not None:
            await on_progress(processed)
    return imported, skips


# --- Academic-year auto-derivation (design §4 step 1) ----------------------
# Default the intake year from the import date; an explicit Academic_year column
# overrides it per row. Coaching sessions start ~June, so before June counts as
# the prior session (Jun 2026 -> 2026-27, Feb 2026 -> 2025-26).
_AY_START_MONTH = 6


def _derive_current_ay_start(today: date | None = None) -> int:
    today = today or date.today()
    return today.year if today.month >= _AY_START_MONTH else today.year - 1


def _ay_name(start_year: int) -> str:
    return f"{start_year}-{(start_year + 1) % 100:02d}"


def _required_ay_starts(
    target: str | None,
    overrides: dict[str, str] | None,
    courses_by_code: dict[str, Any],
    default_start: int,
) -> set[int]:
    """The academic-year start_years a row needs: its intake year (override or
    the date-derived default) plus the following years its programme spans.
    Mirrors create_batch's span resolution (an existing course's duration wins
    over the row's), so every year a batch will touch gets ensured."""
    duration, start_override = _effective_duration_and_start(overrides)
    start = start_override if start_override is not None else default_start
    course_code = _course_code_for(target, duration)
    existing = courses_by_code.get(course_code.lower())
    eff_duration = max(int(existing.duration_years or 1), 1) if existing else duration
    return {start + i for i in range(eff_duration)}


async def _ensure_academic_year(
    session: AsyncSession,
    branch_id: uuid.UUID,
    start_year: int,
    current_user_id: uuid.UUID | None = None,
    import_id: uuid.UUID | None = None,
):
    """Get-or-create the academic year row for ``start_year`` (name 'YYYY-YY').
    A newly-created year is tagged with ``import_id`` so undo can reclaim it."""
    existing = await academic_repository.get_academic_year_by_start_year(
        session, branch_id, start_year
    )
    if existing:
        return existing, False
    ay = await academic_repository.create_academic_year(
        session,
        branch_id=branch_id,
        name=_ay_name(start_year),
        start_year=start_year,
        end_year=start_year + 1,
        created_by=current_user_id,
    )
    if import_id is not None:
        ay.import_id = import_id
    return ay, True


async def preview_import(
    session: AsyncSession,
    file: UploadFile,
    branch_id: uuid.UUID,
    column_map: dict[str, str] | None = None,
) -> dict[str, Any]:
    """Dry-run a student upload: report row issues and, crucially, which
    batch codes already exist vs. are missing, so the admin can create the
    missing ones (with course/exam-date derived from Target) before any
    rows are written."""
    from app.modules.student.services.student_service import (
        _validate_enrolment_fields,
    )

    content = await file.read()
    rows = _apply_column_map(_parse_file(file.filename, content), column_map)
    unrecognized_columns = _unrecognized_columns(rows)

    years = await academic_repository.list_academic_years(session, branch_id)
    academic_year = _pick_academic_year(years)
    placeholder_year_id = academic_year.id if academic_year else uuid.uuid4()
    # Default intake year derived from today; the Academic_year column overrides
    # it per row. We report (but don't create) any years the import would add.
    default_start = _derive_current_ay_start()
    existing_ay_starts = {y.start_year for y in years}
    required_ay_starts: set[int] = {default_start}

    batch_index = await _load_batch_index(session, branch_id)
    courses = await academic_repository.list_courses(session, branch_id)
    courses_by_code = {c.code.strip().lower(): c for c in courses}
    seen_enrols, seen_emails = await _load_student_dedup_keys(session, branch_id)
    seen_fuzzy = await _load_fuzzy_keys(session, branch_id)

    total_rows = 0
    rows_missing_name = 0
    rows_invalid_enrolment = 0
    rows_invalid_consistency = 0
    rows_with_warnings = 0
    rows_possible_duplicate = 0
    duplicate_rows = 0
    unbatched_rows = 0
    importable_rows = 0
    row_issues: list[str] = []

    # Group by normalized code; keep the first-seen display spelling and a
    # tally of Targets so we can derive a course for the missing ones.
    code_display: dict[str, str] = {}
    code_counts: Counter = Counter()
    code_targets: dict[str, Counter] = {}
    # First-seen Course_opt / Duration / Academic_year override per batch code,
    # plus the full set of distinct values per field so disagreements surface.
    code_overrides: dict[str, dict[str, str]] = {}
    code_override_values: dict[str, dict[str, set[str]]] = {}
    # Importable rows per code (valid + non-duplicate) so blocked batches can
    # be discounted from importable_rows once blockers are known.
    code_ok_counts: Counter = Counter()

    for idx, row in enumerate(rows, start=2):  # row 1 is header
        rownum = _row_number(row, idx)
        total_rows += 1
        kwargs = _row_to_student_kwargs(row, branch_id, placeholder_year_id)
        if kwargs is None:
            rows_missing_name += 1
            if len(row_issues) < 20:
                row_issues.append(f"Row {rownum}: missing required 'Name'")
            continue

        target = kwargs.get("target_exam")
        try:
            _validate_enrolment_fields(kwargs.get("standard"), target)
            enrolment_ok = True
        except HTTPException as exc:
            enrolment_ok = False
            rows_invalid_enrolment += 1
            if len(row_issues) < 20:
                row_issues.append(f"Row {rownum}: {exc.detail}")

        # §3 cross-field checks: contradictions block the row; warnings pass.
        row_ov = _row_overrides(row)
        c_errors, c_warnings = _validate_row_consistency(
            kwargs.get("standard"), target, row_ov
        )
        consistency_ok = not c_errors
        if c_errors:
            rows_invalid_consistency += 1
            if len(row_issues) < 20:
                row_issues.append(f"Row {rownum}: {'; '.join(c_errors)}")
        if c_warnings:
            rows_with_warnings += 1
            if len(row_issues) < 20:
                row_issues.append(f"Row {rownum}: warning — {'; '.join(c_warnings)}")

        enr_key, em_key = _dup_keys(kwargs)
        is_dup = (enr_key is not None and enr_key in seen_enrols) or (
            em_key is not None and em_key in seen_emails
        )
        if is_dup:
            duplicate_rows += 1
            if len(row_issues) < 20:
                row_issues.append(
                    f"Row {rownum}: duplicate student (already on file) — will be skipped"
                )
        else:
            if enr_key is not None:
                seen_enrols.add(enr_key)
            if em_key is not None:
                seen_emails.add(em_key)
            # Fuzzy possible-duplicate (name+phone / name+DOB): a soft warning,
            # not a skip — the row still imports.
            fkeys = _fuzzy_keys(kwargs)
            if fkeys & seen_fuzzy:
                rows_possible_duplicate += 1
                if len(row_issues) < 20:
                    row_issues.append(
                        f"Row {rownum}: possible duplicate "
                        f"(same name + phone/DOB) — imported with a warning"
                    )
            seen_fuzzy |= fkeys

        # A row only counts as importable if it is valid, consistent, AND not
        # a duplicate.
        will_import = enrolment_ok and consistency_ok and not is_dup

        if will_import:
            # Every year this row's programme will touch (intake + span) gets
            # ensured at import; collect them so the preview can report new ones.
            required_ay_starts |= _required_ay_starts(
                target, row_ov, courses_by_code, default_start
            )

        code = kwargs.get("_batch_code")
        if code:
            norm = _norm_code(str(code))
            code_display.setdefault(norm, str(code).strip())
            code_counts[norm] += 1
            code_targets.setdefault(norm, Counter())
            # Only rows that will actually be imported drive the batch's derived
            # course / overrides — a row we're going to skip (bad enrolment, a
            # §3 contradiction, or a duplicate) must not skew the Target or
            # fabricate an override conflict that blocks the valid rows.
            if will_import:
                if target:
                    code_targets[norm][target] += 1
                _merge_overrides(code_overrides.setdefault(norm, {}), row_ov)
                _track_override_values(
                    code_override_values.setdefault(norm, {}), row_ov
                )
                code_ok_counts[norm] += 1
        else:
            unbatched_rows += 1

        if will_import:
            importable_rows += 1

    # Years we'll auto-create count as available, so a batch isn't falsely
    # flagged as blocked for an academic year the import itself will add.
    ay_start_years = existing_ay_starts | required_ay_starts

    batches: list[dict[str, Any]] = []
    existing = 0
    blocked = 0
    for norm, display in code_display.items():
        exists = norm in batch_index
        if exists:
            existing += 1
        target = _dominant_target(code_targets.get(norm, Counter()))
        ov = code_overrides.get(norm, {})
        duration, start_override = _effective_duration_and_start(ov)
        course_code = _course_code_for(target, duration)
        course_name = _course_name_for(target, duration, ov.get("course_opt"))
        # Exam happens in the calendar year the programme ends.
        eff_start = start_override if start_override is not None else default_start
        exam_end_year = eff_start + duration
        # For a missing batch, check up front whether auto-create would fail,
        # so the admin sees the blocker before committing rather than after.
        # A code whose rows disagree on the overrides is itself a blocker.
        blocker = None
        if not exists:
            blocker = _override_conflict(
                code_override_values.get(norm, {})
            ) or _missing_batch_blocker(
                target, default_start, courses_by_code, ay_start_years, ov
            )
        if blocker:
            blocked += 1
            # Rows pointing at a batch that cannot be created will be skipped,
            # so they are not actually importable.
            importable_rows -= code_ok_counts.get(norm, 0)
        batches.append(
            {
                "code": display,
                "student_count": code_counts[norm],
                "exists": exists,
                "target": target,
                "suggested_course_code": None if exists else course_code,
                "suggested_course_name": None if exists else course_name,
                "suggested_exam_date": (
                    None if exists else _suggested_exam_date(target, exam_end_year)
                ),
                "creatable": exists or blocker is None,
                "blocker": blocker,
            }
        )

    # Missing-and-blocked first (need attention), then missing, then existing;
    # within each, by how many students depend on the code.
    batches.sort(
        key=lambda b: (b["exists"], b["creatable"], -b["student_count"])
    )

    # Academic years the import would auto-create (date-derived default and any
    # from the Academic_year column) — shown so the admin can spot a typo'd year
    # before committing. The import no longer hard-fails on an empty branch.
    new_academic_years = sorted(
        _ay_name(s) for s in (required_ay_starts - existing_ay_starts)
    )

    return {
        "total_rows": total_rows,
        "importable_rows": max(importable_rows, 0),
        "rows_missing_name": rows_missing_name,
        "rows_invalid_enrolment": rows_invalid_enrolment,
        "rows_invalid_consistency": rows_invalid_consistency,
        "rows_with_warnings": rows_with_warnings,
        "rows_possible_duplicate": rows_possible_duplicate,
        "duplicate_rows": duplicate_rows,
        "unbatched_rows": unbatched_rows,
        "existing_batches": existing,
        "missing_batches": len(code_display) - existing,
        "blocked_batches": blocked,
        "blocking_error": None,
        "new_academic_years": new_academic_years,
        "batches": batches,
        "row_issues": row_issues,
        "unrecognized_columns": unrecognized_columns,
    }


async def validate_import_rows(
    session: AsyncSession,
    branch_id: uuid.UUID,
    value_dicts: list[dict[str, Any]],
    create_missing_batches: bool,
) -> list[dict[str, Any]]:
    """Validate a set of (edited) grid rows against the same rules the import
    enforces — without writing anything. Each ``value_dicts`` entry is keyed by
    canonical field key. Returns one ``{index, errors, warnings}`` per row so the
    grid can flag what's wrong and the admin can fix it inline before committing."""
    from app.modules.student.services.student_service import (
        _validate_enrolment_fields,
    )

    years = await academic_repository.list_academic_years(session, branch_id)
    academic_year = _pick_academic_year(years)
    placeholder = academic_year.id if academic_year else uuid.uuid4()
    default_start = _derive_current_ay_start()
    existing_ay_starts = {y.start_year for y in years}
    batch_index = await _load_batch_index(session, branch_id)
    courses = await academic_repository.list_courses(session, branch_id)
    courses_by_code = {c.code.strip().lower(): c for c in courses}
    seen_enrols, seen_emails = await _load_student_dedup_keys(session, branch_id)
    seen_fuzzy = await _load_fuzzy_keys(session, branch_id)

    results: list[dict[str, Any]] = []
    for idx, values in enumerate(value_dicts):
        row = dict(values)
        errors: list[str] = []
        warnings: list[str] = []

        kwargs = _row_to_student_kwargs(row, branch_id, placeholder)
        if kwargs is None:
            results.append(
                {"index": idx, "errors": ["Name is required"], "warnings": []}
            )
            continue

        batch_code = kwargs.pop("_batch_code", None)
        try:
            _validate_enrolment_fields(
                kwargs.get("standard"), kwargs.get("target_exam")
            )
        except HTTPException as exc:
            errors.append(str(exc.detail))

        row_ov = _row_overrides(row)
        c_errors, c_warnings = _validate_row_consistency(
            kwargs.get("standard"), kwargs.get("target_exam"), row_ov
        )
        errors.extend(c_errors)
        warnings.extend(c_warnings)

        enr_key, em_key = _dup_keys(kwargs)
        is_dup = (enr_key is not None and enr_key in seen_enrols) or (
            em_key is not None and em_key in seen_emails
        )
        if is_dup:
            errors.append("Duplicate student (already on file)")
        else:
            if enr_key is not None:
                seen_enrols.add(enr_key)
            if em_key is not None:
                seen_emails.add(em_key)

        if batch_code:
            norm = _norm_code(str(batch_code))
            if norm not in batch_index:
                if not create_missing_batches:
                    errors.append(f"Unknown batch '{batch_code}'")
                else:
                    req = _required_ay_starts(
                        kwargs.get("target_exam"),
                        row_ov,
                        courses_by_code,
                        default_start,
                    )
                    blocker = _missing_batch_blocker(
                        kwargs.get("target_exam"),
                        default_start,
                        courses_by_code,
                        existing_ay_starts | req | {default_start},
                        row_ov,
                    )
                    if blocker:
                        errors.append(
                            f"Batch '{batch_code}' can't be created — {blocker}"
                        )
                    else:
                        warnings.append(f"Batch '{batch_code}' will be created")

        # Typed-field shape warnings (don't block; import leaves them blank).
        for field, label in (
            ("date_of_birth", "Date of Birth"),
            ("enrollment_date", "Enrollment Date"),
        ):
            v = kwargs.get(field)
            if v and _parse_date(v) is None:
                warnings.append(f"Unrecognized {label} — will be left blank")
        fees = kwargs.get("fees_status")
        if fees and str(fees).strip().lower() not in _VALID_FEES:
            warnings.append("Unrecognized Fees status — will be left blank")

        if not is_dup:
            fkeys = _fuzzy_keys(kwargs)
            if fkeys & seen_fuzzy:
                warnings.append("Possible duplicate (same name + phone/DOB)")
            seen_fuzzy |= fkeys

        results.append({"index": idx, "errors": errors, "warnings": warnings})

    return results


async def import_students(
    session: AsyncSession,
    content: bytes,
    filename: str | None,
    branch_id: uuid.UUID,
    current_user_id: uuid.UUID,
    create_missing_batches: bool = False,
    ip_address: str | None = None,
    import_id: uuid.UUID | None = None,
    on_progress=None,
    column_map: dict[str, str] | None = None,
) -> dict[str, Any]:
    """Engine for a student import. ``content``/``filename`` are the raw upload
    (read by the caller, so this works both in-request and from a background
    job). When ``on_progress`` is given it's awaited with the running count of
    inserted students so a job can publish a progress bar. ``import_id`` lets a
    job key the run to its own id; otherwise one is generated. ``column_map``
    (from the mapping step) renames file headers to canonical fields first."""
    from app.modules.student.services.student_service import (
        VALID_STREAMS,
        _validate_enrolment_fields,
        default_stream,
    )

    rows = _apply_column_map(_parse_file(filename, content), column_map)

    # One id per import run, stamped on every student (and auto-created batch)
    # so this upload can be audited and undone as a unit (design §9).
    if import_id is None:
        import_id = uuid.uuid4()
    source_file = filename

    # Auto-derive the intake year from the import date (overridable per row by
    # the Academic_year column) and ensure every year the import will touch
    # exists — no "create the year first" prerequisite (design §4 step 1).
    default_start = _derive_current_ay_start()
    years = await academic_repository.list_academic_years(session, branch_id)
    courses = await academic_repository.list_courses(session, branch_id)
    courses_by_code = {c.code.strip().lower(): c for c in courses}

    required_starts: set[int] = {default_start}
    for row in rows:
        kwargs = _row_to_student_kwargs(row, branch_id, uuid.uuid4())
        if kwargs is None:
            continue
        try:
            _validate_enrolment_fields(
                kwargs.get("standard"), kwargs.get("target_exam")
            )
        except HTTPException:
            continue
        row_ov = _row_overrides(row)
        c_errors, _ = _validate_row_consistency(
            kwargs.get("standard"), kwargs.get("target_exam"), row_ov
        )
        if c_errors:
            continue
        required_starts |= _required_ay_starts(
            kwargs.get("target_exam"), row_ov, courses_by_code, default_start
        )

    academic_years_created: list[str] = []
    for start in sorted(required_starts):
        _, created = await _ensure_academic_year(
            session, branch_id, start, current_user_id, import_id=import_id
        )
        if created:
            academic_years_created.append(_ay_name(start))

    years = await academic_repository.list_academic_years(session, branch_id)
    academic_year = _ay_by_start_year(years, default_start) or _pick_academic_year(
        years
    )
    academic_year_id = academic_year.id

    batch_index = await _load_batch_index(session, branch_id)
    batches_created: list[str] = []
    subjects_created = 0
    # norm code -> human reason a missing batch could NOT be created, so rows
    # pointing at it report *why* instead of a misleading "unknown batch code".
    batch_create_errors: dict[str, str] = {}
    # Seeded with existing students so a re-upload (or within-file repeats) is
    # skipped instead of silently duplicating every row.
    seen_enrols, seen_emails = await _load_student_dedup_keys(session, branch_id)
    # Soft (name+phone / name+DOB) keys for the fuzzy possible-duplicate warning.
    seen_fuzzy = await _load_fuzzy_keys(session, branch_id)

    # Optionally create any referenced-but-missing batches up front, so the
    # course/exam-date is decided once per code from its dominant Target.
    if create_missing_batches:
        missing: dict[str, tuple[str, Counter]] = {}
        # First-seen override per code, plus the distinct values per field so a
        # code whose rows disagree can be rejected instead of silently guessed.
        code_overrides: dict[str, dict[str, str]] = {}
        code_override_values: dict[str, dict[str, set[str]]] = {}
        # Mirror the main loop's skip logic (invalid enrolment, §3 contradiction,
        # duplicate) so a row we're going to skip never drives a batch's derived
        # course or fabricates an override conflict that blocks the valid rows.
        seen_e, seen_m = set(seen_enrols), set(seen_emails)
        for row in rows:
            kwargs = _row_to_student_kwargs(row, branch_id, academic_year_id)
            if kwargs is None:
                continue
            code = kwargs.get("_batch_code")
            if not code:
                continue
            try:
                _validate_enrolment_fields(
                    kwargs.get("standard"), kwargs.get("target_exam")
                )
            except HTTPException:
                continue
            row_ov = _row_overrides(row)
            c_errors, _ = _validate_row_consistency(
                kwargs.get("standard"), kwargs.get("target_exam"), row_ov
            )
            if c_errors:
                continue
            enr_key, em_key = _dup_keys(kwargs)
            if (enr_key is not None and enr_key in seen_e) or (
                em_key is not None and em_key in seen_m
            ):
                continue
            if enr_key is not None:
                seen_e.add(enr_key)
            if em_key is not None:
                seen_m.add(em_key)
            norm = _norm_code(str(code))
            _merge_overrides(code_overrides.setdefault(norm, {}), row_ov)
            _track_override_values(
                code_override_values.setdefault(norm, {}), row_ov
            )
            if norm in batch_index:
                continue
            display, counter = missing.setdefault(
                norm, (str(code).strip(), Counter())
            )
            target = kwargs.get("target_exam")
            if target:
                counter[target] += 1

        for norm, (display, counter) in missing.items():
            # Rows that disagree on this code's overrides are a hard error —
            # don't guess; rows fall through to "couldn't create batch — …".
            conflict = _override_conflict(code_override_values.get(norm, {}))
            if conflict:
                batch_create_errors[norm] = conflict
                continue
            # Each creation runs in its own savepoint: a failure (e.g. a
            # missing academic year for the derived course's span) rolls back
            # just this batch + any course it created, leaving the outer
            # import transaction usable for the remaining rows.
            try:
                async with session.begin_nested():
                    batch, subj_n = await _create_derived_batch(
                        session,
                        branch_id,
                        display,
                        _dominant_target(counter),
                        academic_year,
                        current_user_id,
                        ip_address,
                        import_id=import_id,
                        overrides=code_overrides.get(norm),
                        years=years,
                    )
                batch_index[norm] = batch.id
                batches_created.append(display)
                subjects_created += subj_n
            except HTTPException as exc:
                batch_create_errors[norm] = str(exc.detail)
            except Exception as exc:  # noqa: BLE001
                batch_create_errors[norm] = str(exc)

    imported = 0
    skipped = 0
    errors: list[str] = []
    warnings: list[str] = []
    # Validated rows are collected here and written in bulk after the loop —
    # per-row INSERT+flush+savepoint is what made a 1000-row import take ~40s
    # and time out. (rownum, Student, StudentBatchMapping | None, Parent | None)
    pending: list[
        tuple[int, Student, StudentBatchMapping | None, Parent | None]
    ] = []

    for idx, row in enumerate(rows, start=2):  # row 1 is header
        rownum = _row_number(row, idx)
        kwargs = _row_to_student_kwargs(row, branch_id, academic_year_id)
        if kwargs is None:
            skipped += 1
            errors.append(f"Row {rownum}: missing required 'Name'")
            continue

        batch_code = kwargs.pop("_batch_code", None)
        try:
            _validate_enrolment_fields(
                kwargs.get("standard"), kwargs.get("target_exam")
            )
        except HTTPException as exc:
            skipped += 1
            errors.append(f"Row {rownum}: {exc.detail}")
            continue

        # §3 cross-field checks: contradictions skip the row; warnings pass.
        c_errors, c_warnings = _validate_row_consistency(
            kwargs.get("standard"), kwargs.get("target_exam"), _row_overrides(row)
        )
        if c_errors:
            skipped += 1
            errors.append(f"Row {rownum}: {'; '.join(c_errors)}")
            continue
        warnings.extend(f"Row {rownum}: {w}" for w in c_warnings)

        enr_key, em_key = _dup_keys(kwargs)
        if (enr_key is not None and enr_key in seen_enrols) or (
            em_key is not None and em_key in seen_emails
        ):
            skipped += 1
            errors.append(
                f"Row {rownum}: duplicate student (already on file) — skipped"
            )
            continue

        batch_id: uuid.UUID | None = None
        if batch_code:
            norm = _norm_code(str(batch_code))
            batch_id = batch_index.get(norm)
            if batch_id is None:
                skipped += 1
                reason = batch_create_errors.get(norm)
                if reason:
                    errors.append(
                        f"Row {rownum}: couldn't create batch '{batch_code}' — {reason}"
                    )
                else:
                    errors.append(f"Row {rownum}: unknown batch code '{batch_code}'")
                continue

        # Resolve the per-student stream: an explicit (valid) Stream value wins;
        # blank or unrecognized falls back to the Target default (PCB for
        # MHT-CET). Admins can override later on the student record.
        raw_stream = kwargs.get("stream")
        if raw_stream:
            norm_stream = str(raw_stream).strip().upper()
            if norm_stream in VALID_STREAMS:
                kwargs["stream"] = norm_stream
            else:
                warnings.append(
                    f"Row {rownum}: unrecognized Stream '{raw_stream}' — "
                    f"defaulted from Target"
                )
                kwargs["stream"] = default_stream(kwargs.get("target_exam"))
        else:
            kwargs["stream"] = default_stream(kwargs.get("target_exam"))

        # Coerce optional typed fields (DOB → date, fees → enum). An unparseable
        # value is dropped with a warning so the rest of the row still imports.
        warnings.extend(_coerce_typed_fields(kwargs, rownum))

        # Fuzzy possible-duplicate: same name + phone/DOB as a student already on
        # file (or an earlier row) but a different/absent roll number. Warn only —
        # never skip, since homonyms are real.
        fkeys = _fuzzy_keys(kwargs)
        if fkeys & seen_fuzzy:
            warnings.append(
                f"Row {rownum}: possible duplicate of an existing student "
                f"(same name + phone/DOB) — imported anyway"
            )
        seen_fuzzy |= fkeys

        # Build the row in memory (id pre-generated so the batch mapping can
        # reference it) and defer the write to a single bulk insert below.
        sid = uuid.uuid4()
        student = Student(
            id=sid,
            import_id=import_id,
            import_source_file=source_file,
            **kwargs,
        )
        mapping = (
            StudentBatchMapping(
                student_id=sid, batch_id=batch_id, branch_id=branch_id
            )
            if batch_id is not None
            else None
        )
        parent = _build_parent(
            row, sid, branch_id, kwargs.get("parent_mobile"), import_id
        )
        pending.append((rownum, student, mapping, parent))
        # Mark the natural keys now so a later row that repeats them within this
        # same file is caught as a duplicate.
        if enr_key is not None:
            seen_enrols.add(enr_key)
        if em_key is not None:
            seen_emails.add(em_key)

    # Bulk-write the validated rows in chunks (one flush per chunk instead of
    # per row). A chunk whose flush fails (e.g. a stray IntegrityError) falls
    # back to per-row savepoints so one bad row can't sink the whole chunk.
    imported, chunk_skips = await _bulk_insert_students(
        session, pending, on_progress=on_progress
    )
    for rn, exc in chunk_skips:
        skipped += 1
        errors.append(f"Row {rn}: {exc}")

    await audit_service.log_action(
        session,
        user_id=current_user_id,
        action="IMPORT",
        table_name="students",
        record_id=import_id,
        new_values={
            "import_id": str(import_id),
            "source_file": source_file,
            "imported": imported,
            "skipped": skipped,
            "batches_created": batches_created,
            "subjects_created": subjects_created,
            "academic_years_created": academic_years_created,
        },
        ip_address=ip_address,
        branch_id=branch_id,
    )

    return {
        "imported": imported,
        "skipped": skipped,
        "errors": errors,
        "warnings": warnings,
        "batches_created": batches_created,
        "subjects_created": subjects_created,
        "academic_years_created": academic_years_created,
        # Only hand back an undo handle when rows actually persisted.
        "import_id": import_id if imported > 0 else None,
    }


async def start_import_job(
    session: AsyncSession,
    content: bytes,
    filename: str | None,
    branch_id: uuid.UUID,
    current_user_id: uuid.UUID,
    create_missing_batches: bool,
) -> StudentImportJob:
    """Create a pending import-job row (with the row count for the progress bar)
    and return it. The heavy work runs afterwards in ``run_import_job`` so the
    request returns immediately — no timeout regardless of file size."""
    try:
        total = len(_parse_file(filename, content))
    except HTTPException:
        total = 0
    job = StudentImportJob(
        branch_id=branch_id,
        created_by=current_user_id,
        filename=filename,
        job_status="pending",
        create_missing_batches=create_missing_batches,
        total_rows=total,
    )
    session.add(job)
    # Commit now (not just flush): the background task opens its OWN session and
    # must find this row. FastAPI runs background tasks before the request
    # session's commit, so without an explicit commit here the job wouldn't be
    # visible yet and run_import_job would no-op (job stuck at 0%).
    await session.commit()
    return job


async def run_import_job(
    job_id: uuid.UUID,
    content: bytes,
    filename: str | None,
    branch_id: uuid.UUID,
    current_user_id: uuid.UUID,
    create_missing_batches: bool,
    ip_address: str | None = None,
    column_map: dict[str, str] | None = None,
) -> None:
    """Background entrypoint: opens its own session (the request's is gone) and
    delegates to _process_import_job."""
    from app.core.database.session import async_session_factory

    async with async_session_factory() as session:
        await _process_import_job(
            session,
            job_id,
            content,
            filename,
            branch_id,
            current_user_id,
            create_missing_batches,
            ip_address,
            column_map,
        )


async def _process_import_job(
    session: AsyncSession,
    job_id: uuid.UUID,
    content: bytes,
    filename: str | None,
    branch_id: uuid.UUID,
    current_user_id: uuid.UUID,
    create_missing_batches: bool,
    ip_address: str | None = None,
    column_map: dict[str, str] | None = None,
) -> None:
    """Run the import keyed to ``job_id`` (= import_id), publishing progress per
    chunk and recording the final result or failure on the job row."""
    # The job is committed by the request before this runs, but retry briefly in
    # case the background task wins a timing race against the commit landing.
    job = await session.get(StudentImportJob, job_id)
    for _ in range(5):
        if job is not None:
            break
        await asyncio.sleep(0.3)
        job = await session.get(StudentImportJob, job_id)
    if job is None:
        return
    job.job_status = "processing"
    await session.commit()

    async def _progress(done: int) -> None:
        job.processed_rows = done
        await session.commit()

    try:
        result = await import_students(
            session,
            content,
            filename,
            branch_id,
            current_user_id,
            create_missing_batches=create_missing_batches,
            ip_address=ip_address,
            import_id=job_id,
            on_progress=_progress,
            column_map=column_map,
        )
        job.imported = result["imported"]
        job.skipped = result["skipped"]
        job.subjects_created = result["subjects_created"]
        job.errors = result["errors"]
        job.warnings = result["warnings"]
        job.batches_created = result["batches_created"]
        job.academic_years_created = result["academic_years_created"]
        job.processed_rows = job.total_rows
        job.job_status = "completed"
        await session.commit()
    except Exception as exc:  # noqa: BLE001 - record failure on the job
        await session.rollback()
        job = await session.get(StudentImportJob, job_id)
        if job is not None:
            job.job_status = "failed"
            job.error_detail = str(exc)[:1000]
            await session.commit()


async def undo_import(
    session: AsyncSession,
    import_id: uuid.UUID,
    branch_id: uuid.UUID,
    current_user_id: uuid.UUID,
    ip_address: str | None = None,
) -> dict[str, Any]:
    """Reverse a bulk import: soft-delete the students it created (and their
    batch mappings), then soft-delete any batches that import auto-created
    which have no other students left, and any subject skeleton it created on
    which no chapters have since been loaded. Scoped to the branch so an import
    id can't reach across branches. Idempotent — re-running finds nothing."""
    from app.modules.academic.models.academic_models import (
        AcademicYear,
        Chapter,
        Course,
        Subject,
        Topic,
    )
    students = (
        await session.execute(
            select(Student).where(
                Student.import_id == import_id,
                Student.branch_id == branch_id,
                Student.is_deleted == False,  # noqa: E712
            )
        )
    ).scalars().all()
    student_ids = [s.id for s in students]

    mappings = []
    if student_ids:
        mappings = (
            await session.execute(
                select(StudentBatchMapping).where(
                    StudentBatchMapping.student_id.in_(student_ids),
                    StudentBatchMapping.is_deleted == False,  # noqa: E712
                )
            )
        ).scalars().all()

    # Parent rows this import created (stamped with the same import_id).
    parents = (
        await session.execute(
            select(Parent).where(
                Parent.import_id == import_id,
                Parent.branch_id == branch_id,
                Parent.is_deleted == False,  # noqa: E712
            )
        )
    ).scalars().all()

    for m in mappings:
        m.is_deleted = True
    for p in parents:
        p.is_deleted = True
    for s in students:
        s.is_deleted = True
    await session.flush()

    # Batches this import spun up. Only reclaim ones with no remaining (live)
    # student — a batch someone has since assigned other students to is kept.
    batches = (
        await session.execute(
            select(Batch).where(
                Batch.import_id == import_id,
                Batch.branch_id == branch_id,
                Batch.is_deleted == False,  # noqa: E712
            )
        )
    ).scalars().all()

    batches_deleted = 0
    for b in batches:
        remaining = (
            await session.execute(
                select(func.count())
                .select_from(StudentBatchMapping)
                .where(
                    StudentBatchMapping.batch_id == b.id,
                    StudentBatchMapping.is_deleted == False,  # noqa: E712
                )
            )
        ).scalar_one()
        if remaining == 0:
            b.is_deleted = True
            batches_deleted += 1
    await session.flush()

    # Curriculum this import auto-populated: remove exactly what it created —
    # its topics, then its chapters. A syllabus import (or the academic UI) that
    # added its own chapters/topics leaves those untouched (their import_id is
    # NULL / a different id).
    for table in (Topic, Chapter):
        rows = (
            await session.execute(
                select(table).where(
                    table.import_id == import_id,
                    table.branch_id == branch_id,
                    table.is_deleted == False,  # noqa: E712
                )
            )
        ).scalars().all()
        for r in rows:
            r.is_deleted = True
    await session.flush()

    # Reclaim each subject this import created that now has no chapters left —
    # i.e. nothing but the (now-removed) skeleton. A subject a syllabus import
    # has since built its own chapters on is kept (§7.4).
    subjects = (
        await session.execute(
            select(Subject).where(
                Subject.import_id == import_id,
                Subject.branch_id == branch_id,
                Subject.is_deleted == False,  # noqa: E712
            )
        )
    ).scalars().all()

    subjects_deleted = 0
    for s in subjects:
        remaining_chapters = (
            await session.execute(
                select(func.count())
                .select_from(Chapter)
                .where(
                    Chapter.subject_id == s.id,
                    Chapter.is_deleted == False,  # noqa: E712
                )
            )
        ).scalar_one()
        if remaining_chapters == 0:
            s.is_deleted = True
            subjects_deleted += 1
    await session.flush()

    # Reclaim courses this import created, but only when nothing live still
    # points at them — no live batch on the course and no live subject under it
    # (a course a syllabus import has since built curriculum on is kept).
    courses = (
        await session.execute(
            select(Course).where(
                Course.import_id == import_id,
                Course.branch_id == branch_id,
                Course.is_deleted == False,  # noqa: E712
            )
        )
    ).scalars().all()

    courses_deleted = 0
    for c in courses:
        live_batches = (
            await session.execute(
                select(func.count())
                .select_from(Batch)
                .where(
                    Batch.course_id == c.id,
                    Batch.is_deleted == False,  # noqa: E712
                )
            )
        ).scalar_one()
        live_subjects = (
            await session.execute(
                select(func.count())
                .select_from(Subject)
                .where(
                    Subject.course_id == c.id,
                    Subject.is_deleted == False,  # noqa: E712
                )
            )
        ).scalar_one()
        if live_batches == 0 and live_subjects == 0:
            c.is_deleted = True
            courses_deleted += 1
    await session.flush()

    # Reclaim academic years this import created, but only when no live batch
    # (start or end), subject, or student still references them.
    academic_years = (
        await session.execute(
            select(AcademicYear).where(
                AcademicYear.import_id == import_id,
                AcademicYear.branch_id == branch_id,
                AcademicYear.is_deleted == False,  # noqa: E712
            )
        )
    ).scalars().all()

    academic_years_deleted = 0
    for ay in academic_years:
        refs = (
            await session.execute(
                select(func.count())
                .select_from(Batch)
                .where(
                    Batch.is_deleted == False,  # noqa: E712
                    (Batch.start_academic_year_id == ay.id)
                    | (Batch.end_academic_year_id == ay.id),
                )
            )
        ).scalar_one()
        refs += (
            await session.execute(
                select(func.count())
                .select_from(Subject)
                .where(
                    Subject.academic_year_id == ay.id,
                    Subject.is_deleted == False,  # noqa: E712
                )
            )
        ).scalar_one()
        refs += (
            await session.execute(
                select(func.count())
                .select_from(Student)
                .where(
                    Student.academic_year_id == ay.id,
                    Student.is_deleted == False,  # noqa: E712
                )
            )
        ).scalar_one()
        if refs == 0:
            ay.is_deleted = True
            academic_years_deleted += 1
    await session.flush()

    await audit_service.log_action(
        session,
        user_id=current_user_id,
        action="IMPORT_UNDO",
        table_name="students",
        record_id=import_id,
        new_values={
            "import_id": str(import_id),
            "students_deleted": len(students),
            "batches_deleted": batches_deleted,
            "subjects_deleted": subjects_deleted,
            "parents_deleted": len(parents),
            "courses_deleted": courses_deleted,
            "academic_years_deleted": academic_years_deleted,
        },
        ip_address=ip_address,
        branch_id=branch_id,
    )

    return {
        "students_deleted": len(students),
        "batches_deleted": batches_deleted,
        "subjects_deleted": subjects_deleted,
        "parents_deleted": len(parents),
        "courses_deleted": courses_deleted,
        "academic_years_deleted": academic_years_deleted,
    }
