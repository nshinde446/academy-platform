"""Advanced Teacher Productivity report exporters — Excel (openpyxl) and PDF
(headless Chromium, reusing the attendance export engine).

Pure builders: they take the already-assembled report dict from
``lecture_service.get_productivity_report`` and return file bytes. No DB work.
"""

from __future__ import annotations

import html
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Reuse the same headless-Chromium HTML->PDF engine as the attendance reports.
from app.modules.attendance.services.attendance_export_service import (
    render_html_to_pdf,
)

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
PDF_MIME = "application/pdf"

_HEAD_FILL = PatternFill("solid", fgColor="34558B")
_HEAD_FONT = Font(bold=True, color="FFFFFF")
_BOLD = Font(bold=True)
_CENTER = Alignment(horizontal="center")


def _pct(v: float | None) -> str:
    return "—" if v is None else f"{v}%"


def _period(from_date, to_date) -> str:
    a = from_date.date().isoformat() if from_date else "start"
    b = to_date.date().isoformat() if to_date else "now"
    return f"{a} to {b}"


# ── Excel ────────────────────────────────────────────────────────────────────


def _write_sheet(ws, headers: list[str], rows: list[list]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = _HEAD_FILL
        cell.font = _HEAD_FONT
        cell.alignment = _CENTER
    for r in rows:
        ws.append(r)
    for i, _ in enumerate(headers, start=1):
        col = get_column_letter(i)
        width = max(
            [len(str(headers[i - 1]))] + [len(str(r[i - 1])) for r in rows] + [8]
        )
        ws.column_dimensions[col].width = min(width + 2, 40)


def productivity_report_xlsx(report: dict, *, brand: str) -> bytes:
    wb = Workbook()

    ws = wb.active
    ws.title = "Teachers"
    _write_sheet(
        ws,
        [
            "Teacher",
            "Scheduled",
            "Conducted",
            "Completion %",
            "Punctuality %",
            "Avg Delay (min)",
            "Hours",
            "Topics Planned",
            "Topics Covered",
        ],
        [
            [
                f"{t['first_name']} {t['last_name']}",
                t["scheduled"],
                t["conducted"],
                _pct(t["completion_pct"]),
                _pct(t["punctuality_pct"]),
                t["avg_delay_min"],
                t["hours"],
                t["topics_planned"],
                t["topics_covered"],
            ]
            for t in report["by_teacher"]
        ],
    )

    ws2 = wb.create_sheet("Subjects")
    _write_sheet(
        ws2,
        ["Subject", "Scheduled", "Conducted", "Completion %", "Hours"],
        [
            [
                s["subject_name"],
                s["scheduled"],
                s["conducted"],
                _pct(s["completion_pct"]),
                s["hours"],
            ]
            for s in report["by_subject"]
        ],
    )

    ws3 = wb.create_sheet("Batches")
    _write_sheet(
        ws3,
        ["Batch", "Scheduled", "Conducted", "Completion %", "Hours"],
        [
            [
                b["batch_name"],
                b["scheduled"],
                b["conducted"],
                _pct(b["completion_pct"]),
                b["hours"],
            ]
            for b in report["by_batch"]
        ],
    )

    ws4 = wb.create_sheet("Weekly Trend")
    _write_sheet(
        ws4,
        ["Week", "Scheduled", "Conducted", "Completion %", "Punctuality %", "Hours"],
        [
            [
                w["label"],
                w["scheduled"],
                w["conducted"],
                _pct(w["completion_pct"]),
                _pct(w["punctuality_pct"]),
                w["hours"],
            ]
            for w in report["trend"]
        ],
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── PDF (HTML -> Chromium) ───────────────────────────────────────────────────

_PDF_CSS = """
* { box-sizing: border-box; }
body { font-family: Arial, Helvetica, sans-serif; color: #1f2733; font-size: 11px; }
h1 { font-size: 20px; margin: 0 0 2px; color: #2f4b82; }
.sub { color: #6b7280; margin: 0 0 14px; font-size: 11px; }
.cards { display: flex; gap: 10px; margin-bottom: 16px; }
.card { flex: 1; border: 1px solid #e2e6ec; border-radius: 8px; padding: 8px 10px; }
.card .lbl { color: #6b7280; font-size: 10px; }
.card .val { font-size: 18px; font-weight: 700; }
h2 { font-size: 13px; margin: 16px 0 6px; color: #2f4b82; }
table { width: 100%; border-collapse: collapse; margin-bottom: 8px; }
th { background: #34558b; color: #fff; text-align: left; padding: 5px 7px; font-size: 10px; }
td { padding: 4px 7px; border-bottom: 1px solid #eef1f5; font-size: 10px; }
td.num, th.num { text-align: right; }
tr:nth-child(even) td { background: #f8fafc; }
"""


def _card(label: str, value: str) -> str:
    return (
        f"<div class='card'><div class='lbl'>{html.escape(label)}</div>"
        f"<div class='val'>{html.escape(value)}</div></div>"
    )


def _table(headers: list[str], rows: list[list[str]], num_from: int) -> str:
    head = "".join(
        f"<th class='{'num' if i >= num_from else ''}'>{html.escape(h)}</th>"
        for i, h in enumerate(headers)
    )
    body = "".join(
        "<tr>"
        + "".join(
            f"<td class='{'num' if i >= num_from else ''}'>{html.escape(str(c))}</td>"
            for i, c in enumerate(r)
        )
        + "</tr>"
        for r in rows
    )
    return f"<table><thead><tr>{head}</tr></thead><tbody>{body}</tbody></table>"


def productivity_report_pdf(report: dict, *, brand: str) -> str:
    """Build the report HTML (the caller renders it via ``render_html_to_pdf``)."""
    s = report["summary"]
    period = _period(report.get("from_date"), report.get("to_date"))

    cards = "".join(
        [
            _card("Total Scheduled", str(s["total_scheduled"])),
            _card("Total Conducted", str(s["total_conducted"])),
            _card("Completion %", _pct(s["completion_pct"])),
            _card("Punctuality %", _pct(s["punctuality_pct"])),
            _card("Total Hours", f"{s['total_hours']}h"),
        ]
    )

    teacher_tbl = _table(
        ["Teacher", "Sched", "Cond", "Compl %", "Punc %", "Avg Delay", "Hours", "Topics"],
        [
            [
                f"{t['first_name']} {t['last_name']}",
                t["scheduled"],
                t["conducted"],
                _pct(t["completion_pct"]),
                _pct(t["punctuality_pct"]),
                f"{t['avg_delay_min']}m",
                f"{t['hours']}h",
                f"{t['topics_covered']}/{t['topics_planned']}",
            ]
            for t in report["by_teacher"]
        ],
        num_from=1,
    )

    subject_tbl = _table(
        ["Subject", "Sched", "Cond", "Compl %", "Hours"],
        [
            [s2["subject_name"], s2["scheduled"], s2["conducted"], _pct(s2["completion_pct"]), f"{s2['hours']}h"]
            for s2 in report["by_subject"]
        ],
        num_from=1,
    )

    batch_tbl = _table(
        ["Batch", "Sched", "Cond", "Compl %", "Hours"],
        [
            [b["batch_name"], b["scheduled"], b["conducted"], _pct(b["completion_pct"]), f"{b['hours']}h"]
            for b in report["by_batch"]
        ],
        num_from=1,
    )

    trend_tbl = _table(
        ["Week", "Sched", "Cond", "Compl %", "Punc %", "Hours"],
        [
            [w["label"], w["scheduled"], w["conducted"], _pct(w["completion_pct"]), _pct(w["punctuality_pct"]), f"{w['hours']}h"]
            for w in report["trend"]
        ],
        num_from=1,
    )

    body = (
        f"<h1>{html.escape(brand)}</h1>"
        f"<p class='sub'>Teacher Productivity Report · {html.escape(period)}</p>"
        f"<div class='cards'>{cards}</div>"
        f"<h2>Teacher-wise</h2>{teacher_tbl}"
        f"<h2>Subject-wise</h2>{subject_tbl}"
        f"<h2>Batch-wise</h2>{batch_tbl}"
        f"<h2>Week-wise trend</h2>{trend_tbl}"
    )
    return (
        "<!DOCTYPE html><html><head><meta charset='utf-8'>"
        f"<style>{_PDF_CSS}</style></head><body>{body}</body></html>"
    )


async def productivity_report_bytes(
    report: dict, *, fmt: str, brand: str
) -> tuple[bytes, str]:
    """Return (bytes, mime) for the requested format ('pdf' | 'xlsx')."""
    if fmt == "xlsx":
        return productivity_report_xlsx(report, brand=brand), XLSX_MIME
    doc = productivity_report_pdf(report, brand=brand)
    return await render_html_to_pdf(doc, landscape=True), PDF_MIME
