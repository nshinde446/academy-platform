"""CSV/Excel bulk import for the weekly lecture schedule.

The institute admin uploads a sheet with columns:

  date            (YYYY-MM-DD)
  start_time      (HH:MM, 24h)
  end_time        (HH:MM, 24h)
  teacher_email   (matches users.email)
  batch_code      (matches batches.code)
  subject_code    (matches subjects.code)
  topic           optional, free-text — ignored if no Topic row matches
  classroom_code  optional, matches classrooms.code
  delivery_mode   optional, "offline" (default) or "online"
  notes           optional

One row → one scheduled Lecture. We reuse schedule_lecture from
lecture_service to keep conflict-detection consistent with single-row
creates. Rows that conflict or reference unknown codes are skipped,
not fatal — the response lists per-row errors so the admin can fix
and re-upload.
"""

from __future__ import annotations

import csv
import io
import uuid
from datetime import datetime, time as _time
from typing import Any

from fastapi import HTTPException, UploadFile, status
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.modules.academic.models.academic_models import Subject
from app.modules.audit.services import audit_service
from app.modules.auth.models.auth_models import User
from app.modules.batch.models.batch_models import Batch
from app.modules.classroom.models.classroom_models import Classroom
from app.modules.lectures.repositories import lecture_repository
from app.modules.lectures.schemas.lecture_schemas import LectureCreate
from app.modules.lectures.services import lecture_service
from app.modules.teacher.models.teacher_models import Teacher
from app.modules.teacher.repositories import teacher_repository


REQUIRED_COLUMNS = (
    "date",
    "start_time",
    "end_time",
    "teacher_email",
    "batch_code",
    "subject_code",
)


def _normalize(header: str) -> str:
    return header.strip().lower().replace(" ", "_")


def _norm_name(value: str) -> str:
    """Lowercase + collapse internal whitespace, so 'SURAJ  SHINDE' and
    'Suraj Shinde' compare equal when matching a teacher by name."""
    return " ".join(str(value).lower().split())


def _parse_csv(content: bytes) -> list[dict[str, Any]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


def _parse_xlsx(content: bytes) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb.active
    if ws is None:
        return []
    rows = ws.iter_rows(values_only=True)
    try:
        headers = [str(h) if h is not None else "" for h in next(rows)]
    except StopIteration:
        return []
    out: list[dict[str, Any]] = []
    for row in rows:
        record = {
            headers[i]: ("" if row[i] is None else row[i])
            for i in range(min(len(headers), len(row)))
        }
        if any(v not in ("", None) for v in record.values()):
            out.append(record)
    return out


def _parse_time(value: Any) -> _time:
    """Accept '09:00', '9:00', '09:00:00', or an openpyxl time/datetime."""
    if isinstance(value, _time):
        return value
    if isinstance(value, datetime):
        return value.time()
    s = str(value).strip()
    for fmt in ("%H:%M", "%H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).time()
        except ValueError:
            continue
    raise ValueError(f"Bad time '{s}' (use HH:MM)")


def _parse_date(value: Any) -> datetime:
    """Accept 'YYYY-MM-DD' or a datetime cell from Excel."""
    if isinstance(value, datetime):
        return datetime(value.year, value.month, value.day)
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    raise ValueError(f"Bad date '{s}' (use YYYY-MM-DD)")


async def _resolve_lookups(
    session: AsyncSession,
    branch_id: uuid.UUID,
    teacher_email: str,
    batch_code: str,
    subject_code: str,
    classroom_code: str | None,
) -> tuple[Teacher, Batch, Subject, Classroom | None]:
    """Resolve human-readable codes to FK ids. Raises ValueError with a
    user-friendly message if anything is missing — caller turns those
    into per-row errors."""
    # Teacher resolution tries two paths. Many teacher rows are linked
    # to a User account (so we look up via Users.email → Teacher.user_id);
    # some teachers — especially seeded / imported ones — have an email
    # stored directly on the Teacher row instead. Either match is fine.
    normalized = teacher_email.lower().strip()
    teacher: Teacher | None = None

    user = (
        await session.execute(select(User).where(User.email == normalized))
    ).scalar_one_or_none()
    if user:
        teacher = (
            await session.execute(
                select(Teacher).where(
                    Teacher.user_id == user.id,
                    Teacher.branch_id == branch_id,
                    Teacher.is_deleted == False,
                )
            )
        ).scalar_one_or_none()
    if teacher is None:
        teacher = (
            await session.execute(
                select(Teacher).where(
                    Teacher.email == normalized,
                    Teacher.branch_id == branch_id,
                    Teacher.is_deleted == False,
                )
            )
        ).scalar_one_or_none()
    if teacher is None:
        # Email isn't always recorded, so also match by full name (first last).
        wanted = _norm_name(teacher_email)
        candidates = (
            await session.execute(
                select(Teacher).where(
                    Teacher.branch_id == branch_id,
                    Teacher.is_deleted == False,
                )
            )
        ).scalars().all()
        matches = [
            t
            for t in candidates
            if _norm_name(f"{t.first_name} {t.last_name}") == wanted
        ]
        if len(matches) == 1:
            teacher = matches[0]
        elif len(matches) > 1:
            raise ValueError(
                f"more than one teacher named '{teacher_email}' — "
                f"use a unique name or email"
            )
    if teacher is None:
        raise ValueError(f"no teacher in this branch matches: {teacher_email}")

    batch = (
        await session.execute(
            select(Batch).where(
                Batch.code == batch_code.strip(),
                Batch.branch_id == branch_id,
                Batch.is_deleted == False,
            )
        )
    ).scalar_one_or_none()
    if not batch:
        raise ValueError(f"unknown batch_code: {batch_code}")

    # Subject.code isn't globally unique (JEE Physics vs NEET Physics), so we
    # disambiguate by the batch's course — the subjects the batch actually
    # studies. (This used to require a BatchSubjectMapping row, but there's no
    # UI/endpoint to create those, so course scoping is both simpler and always
    # available.)
    subject = (
        await session.execute(
            select(Subject).where(
                Subject.code == subject_code.strip(),
                Subject.course_id == batch.course_id,
                Subject.is_deleted == False,
            )
        )
    ).scalar_one_or_none()
    if not subject:
        raise ValueError(
            f"subject '{subject_code}' is not in batch '{batch.code}' course"
        )

    classroom = None
    if classroom_code:
        classroom = (
            await session.execute(
                select(Classroom).where(
                    Classroom.code == classroom_code.strip(),
                    Classroom.branch_id == branch_id,
                    Classroom.is_deleted == False,
                )
            )
        ).scalar_one_or_none()
        if not classroom:
            raise ValueError(f"unknown classroom_code: {classroom_code}")

    return teacher, batch, subject, classroom


async def _holiday_dates(session: AsyncSession, branch_id: uuid.UUID) -> set:
    """The branch's non-teaching days as a set of date, so the importer skips
    rows landing on a holiday — matching generate / copy-to-next-day."""
    rows = await lecture_repository.list_holidays(session, branch_id)
    return {h.holiday_date for h in rows}


async def preview_schedule(
    session: AsyncSession,
    file: UploadFile,
    branch_id: uuid.UUID,
) -> dict[str, Any]:
    """Dry-run validation of an import sheet — resolves every row and reports
    what would happen, WITHOUT creating anything.

    Brings the lecture importer up to the student importer's maturity: the admin
    sees exactly which rows are clean and which would be skipped (and why) before
    committing. Each row is checked for code resolution, time validity, the
    Subject→Teacher lock, teacher leave, and existing-schedule conflicts.
    """
    content = await file.read()
    filename = (file.filename or "").lower()

    if filename.endswith(".csv"):
        rows = _parse_csv(content)
    elif filename.endswith((".xlsx", ".xls")):
        rows = _parse_xlsx(content)
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported file type. Use .csv or .xlsx",
        )

    if rows:
        present = {_normalize(k) for k in rows[0].keys() if k}
        missing = [c for c in REQUIRED_COLUMNS if c not in present]
        if missing:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Missing required columns: {', '.join(missing)}",
            )

    holidays = await _holiday_dates(session, branch_id)
    out_rows: list[dict[str, Any]] = []
    ok_count = 0
    error_count = 0

    for idx, raw in enumerate(rows, start=2):  # row 1 is header
        row = {_normalize(k): v for k, v in raw.items() if k}
        result: dict[str, Any] = {
            "row_number": idx,
            "date": str(row.get("date", "")),
            "start_time": str(row.get("start_time", "")),
            "end_time": str(row.get("end_time", "")),
            "teacher": str(row.get("teacher_email", "")),
            "batch": str(row.get("batch_code", "")),
            "subject": str(row.get("subject_code", "")),
            "status": "ok",
            "message": "",
        }
        try:
            date_part = _parse_date(row.get("date"))
            start_t = _parse_time(row.get("start_time"))
            end_t = _parse_time(row.get("end_time"))
            scheduled_start = datetime.combine(date_part.date(), start_t)
            scheduled_end = datetime.combine(date_part.date(), end_t)
            if scheduled_end <= scheduled_start:
                raise ValueError("end_time must be after start_time")
            if date_part.date() in holidays:
                raise ValueError(f"{date_part.date().isoformat()} is a holiday")

            teacher, batch, subject, classroom = await _resolve_lookups(
                session,
                branch_id,
                str(row.get("teacher_email", "")),
                str(row.get("batch_code", "")),
                str(row.get("subject_code", "")),
                (str(row.get("classroom_code")) if row.get("classroom_code") else None),
            )

            if not await teacher_repository.teacher_teaches_subject(
                session, teacher.id, subject.id
            ):
                raise ValueError(
                    f"{teacher.first_name} {teacher.last_name} isn't assigned to "
                    f"subject '{subject.code}'"
                )
            if await teacher_repository.teacher_on_leave(
                session, teacher.id, date_part.date()
            ):
                raise ValueError("teacher is on leave that day")
            if await lecture_repository.check_teacher_conflict(
                session, teacher.id, scheduled_start, scheduled_end, None
            ):
                raise ValueError("teacher already has a lecture at this time")
            if await lecture_repository.check_batch_conflict(
                session, batch.id, scheduled_start, scheduled_end, None
            ):
                raise ValueError("batch already has a lecture at this time")
            if classroom and await lecture_repository.check_classroom_conflict(
                session, classroom.id, scheduled_start, scheduled_end, None
            ):
                raise ValueError("classroom already booked at this time")

            result["message"] = (
                f"{batch.name} · {subject.name} · "
                f"{teacher.first_name} {teacher.last_name}"
            )
            ok_count += 1
        except Exception as exc:  # noqa: BLE001
            result["status"] = "error"
            result["message"] = str(exc)
            error_count += 1
        out_rows.append(result)

    return {"rows": out_rows, "ok_count": ok_count, "error_count": error_count}


async def import_schedule(
    session: AsyncSession,
    file: UploadFile,
    branch_id: uuid.UUID,
    current_user_id: uuid.UUID,
    ip_address: str | None = None,
) -> dict[str, Any]:
    content = await file.read()
    filename = (file.filename or "").lower()

    if filename.endswith(".csv"):
        rows = _parse_csv(content)
    elif filename.endswith((".xlsx", ".xls")):
        rows = _parse_xlsx(content)
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported file type. Use .csv or .xlsx",
        )

    # Validate required columns up front so admins get one clear error
    # instead of N per-row complaints.
    if rows:
        present = {_normalize(k) for k in rows[0].keys() if k}
        missing = [c for c in REQUIRED_COLUMNS if c not in present]
        if missing:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Missing required columns: {', '.join(missing)}",
            )

    holidays = await _holiday_dates(session, branch_id)
    imported = 0
    skipped = 0
    errors: list[str] = []

    for idx, raw in enumerate(rows, start=2):  # row 1 is header
        row = {_normalize(k): v for k, v in raw.items() if k}

        try:
            date_part = _parse_date(row.get("date"))
            start_t = _parse_time(row.get("start_time"))
            end_t = _parse_time(row.get("end_time"))
            scheduled_start = datetime.combine(date_part.date(), start_t)
            scheduled_end = datetime.combine(date_part.date(), end_t)
            if scheduled_end <= scheduled_start:
                raise ValueError("end_time must be after start_time")
            if date_part.date() in holidays:
                raise ValueError(f"{date_part.date().isoformat()} is a holiday")

            teacher, batch, subject, classroom = await _resolve_lookups(
                session,
                branch_id,
                str(row.get("teacher_email", "")),
                str(row.get("batch_code", "")),
                str(row.get("subject_code", "")),
                (str(row.get("classroom_code")) if row.get("classroom_code") else None),
            )

            # Default delivery mode is offline only when a classroom is
            # supplied; otherwise online (no-room schedules are valid).
            raw_mode = str(row.get("delivery_mode") or "").strip().lower()
            delivery_mode = raw_mode or ("offline" if classroom else "online")

            await lecture_service.schedule_lecture(
                session,
                LectureCreate(
                    teacher_id=teacher.id,
                    batch_id=batch.id,
                    classroom_id=classroom.id if classroom else None,
                    subject_id=subject.id,
                    topic_id=None,
                    scheduled_start=scheduled_start,
                    scheduled_end=scheduled_end,
                    delivery_mode=delivery_mode,
                    notes=(str(row.get("notes")) if row.get("notes") else None),
                ),
                current_user_id,
                ip_address,
            )
            imported += 1
        except HTTPException as exc:
            # Conflict, validation, etc. — surface the message but keep going.
            skipped += 1
            errors.append(f"Row {idx}: {exc.detail}")
        except ValueError as exc:
            skipped += 1
            errors.append(f"Row {idx}: {exc}")
        except Exception as exc:  # noqa: BLE001
            skipped += 1
            errors.append(f"Row {idx}: {exc}")

    await audit_service.log_action(
        session,
        user_id=current_user_id,
        action="IMPORT",
        table_name="lectures",
        record_id=uuid.uuid4(),
        new_values={"imported": imported, "skipped": skipped},
        ip_address=ip_address,
        branch_id=branch_id,
    )

    return {"imported": imported, "skipped": skipped, "errors": errors}
