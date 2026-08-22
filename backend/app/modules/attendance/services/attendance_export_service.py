"""Attendance report builders — Excel (openpyxl) and PDF (headless Chromium).

Pure builders: they take already-queried data (see daily_service) plus a
timezone for local time display, and return file bytes. The DB work lives in
the route. Three report scopes — individual student, single batch (register
matrix), all batches (summary + sheet per batch) — each in both formats.
"""

from __future__ import annotations

import html
import io
from datetime import date, datetime, timezone
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.modules.attendance.time_utils import get_tz

# Cell fills for the register matrix (light, print-friendly).
_FILL = {
    "P": PatternFill("solid", fgColor="D7F0DB"),  # green
    "L": PatternFill("solid", fgColor="FCEFC7"),  # amber
    "A": PatternFill("solid", fgColor="F8D7DA"),  # red
}
_HEAD_FILL = PatternFill("solid", fgColor="EEF1F5")
_BOLD = Font(bold=True)
_CENTER = Alignment(horizontal="center")


def _fmt_time(dt: datetime | None, tz_name: str) -> str:
    if dt is None:
        return "—"
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(get_tz(tz_name)).strftime("%H:%M")


def _period(start: date, end: date) -> str:
    return f"{start.isoformat()} to {end.isoformat()}"


def _xlsx_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Excel ──────────────────────────────────────────────────────────────────


def student_xlsx(
    *, brand: str, student_name: str, start: date, end: date,
    summary: dict, timeline: list[Any], tz_name: str,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    ws["A1"] = brand
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Attendance report — {student_name}"
    ws["A3"] = _period(start, end)
    ws["A4"] = (
        f"Working days {summary['working_days']} · "
        f"Present {summary['present_days']} · "
        f"Absent {summary['absent_days']} · "
        f"{summary['attendance_pct']}%"
    )

    header_row = 6
    headers = ["Date", "In", "Out", "Status", "Sign-off"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.font = _BOLD
        cell.fill = _HEAD_FILL

    rows = sorted(timeline, key=lambda r: r.attendance_date)
    for i, r in enumerate(rows, start=header_row + 1):
        ws.cell(row=i, column=1, value=r.attendance_date.isoformat())
        ws.cell(row=i, column=2, value=_fmt_time(r.first_in, tz_name))
        ws.cell(row=i, column=3, value=_fmt_time(r.last_out, tz_name))
        ws.cell(row=i, column=4, value=r.day_status)
        ws.cell(row=i, column=5, value=r.signoff)

    for c in range(1, 6):
        ws.column_dimensions[get_column_letter(c)].width = 14
    return _xlsx_bytes(wb)


def _write_matrix_sheet(ws, *, title_lines: list[str], matrix: dict) -> None:
    dates: list[date] = matrix["dates"]
    for i, line in enumerate(title_lines, start=1):
        ws.cell(row=i, column=1, value=line)
    ws.cell(row=1, column=1).font = Font(bold=True, size=13)

    head = len(title_lines) + 2
    ws.cell(row=head, column=1, value="#").font = _BOLD
    ws.cell(row=head, column=2, value="Student").font = _BOLD
    ws.cell(row=head, column=3, value="PRN").font = _BOLD
    for j, d in enumerate(dates):
        cell = ws.cell(row=head, column=4 + j, value=f"{d.day}/{d.month}")
        cell.font = _BOLD
        cell.alignment = _CENTER
        cell.fill = _HEAD_FILL
    pct_col = 4 + len(dates)
    ws.cell(row=head, column=pct_col, value="Present").font = _BOLD
    ws.cell(row=head, column=pct_col + 1, value="Absent").font = _BOLD
    ws.cell(row=head, column=pct_col + 2, value="%").font = _BOLD

    for r, srow in enumerate(matrix["students"], start=head + 1):
        ws.cell(row=r, column=1, value=r - head)
        ws.cell(row=r, column=2, value=srow["name"])
        ws.cell(row=r, column=3, value=srow["enrollment_number"] or "")
        for j, code in enumerate(srow["cells"]):
            cell = ws.cell(row=r, column=4 + j, value=code)
            cell.alignment = _CENTER
            if code in _FILL:
                cell.fill = _FILL[code]
        ws.cell(row=r, column=pct_col, value=srow["present"])
        ws.cell(row=r, column=pct_col + 1, value=srow["working_days"] - srow["present"])
        ws.cell(row=r, column=pct_col + 2, value=srow["attendance_pct"])

    totals_row = head + 1 + len(matrix["students"])
    ws.cell(row=totals_row, column=2, value="Present / day").font = _BOLD
    for j, n in enumerate(matrix["day_present"]):
        cell = ws.cell(row=totals_row, column=4 + j, value=n)
        cell.alignment = _CENTER
        cell.font = _BOLD

    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 12
    for j in range(len(dates)):
        ws.column_dimensions[get_column_letter(4 + j)].width = 4.5


def batch_xlsx(
    *, brand: str, batch_name: str, start: date, end: date, matrix: dict,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Register"
    _write_matrix_sheet(
        ws,
        title_lines=[
            f"{brand} — {batch_name}",
            f"Attendance register · {_period(start, end)}",
            f"{matrix['student_count']} students · {len(matrix['dates'])} working days · P present / L late / A absent",
        ],
        matrix=matrix,
    )
    return _xlsx_bytes(wb)


def _safe_sheet_title(name: str, used: set[str]) -> str:
    # Excel sheet titles: <=31 chars, no []:*?/\
    clean = "".join(c for c in name if c not in '[]:*?/\\')[:31] or "Batch"
    base, n = clean, 1
    while clean in used:
        suffix = f" {n}"
        clean = base[: 31 - len(suffix)] + suffix
        n += 1
    used.add(clean)
    return clean


def all_batches_xlsx(
    *, brand: str, start: date, end: date,
    summaries: list[dict], matrices: dict[str, dict], batch_names: dict[str, str],
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"{brand} — all batches"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Attendance summary · {_period(start, end)}"

    head = 4
    cols = ["Batch", "Code", "Students", "Working days", "Present", "Total", "Avg %"]
    for c, h in enumerate(cols, start=1):
        cell = ws.cell(row=head, column=c, value=h)
        cell.font = _BOLD
        cell.fill = _HEAD_FILL
    for i, s in enumerate(summaries, start=head + 1):
        ws.cell(row=i, column=1, value=s["batch_name"])
        ws.cell(row=i, column=2, value=s["batch_code"])
        ws.cell(row=i, column=3, value=s["student_count"])
        ws.cell(row=i, column=4, value=s["working_days"])
        ws.cell(row=i, column=5, value=s["present"])
        ws.cell(row=i, column=6, value=s["total_slots"])
        ws.cell(row=i, column=7, value=s["avg_pct"])
    ws.column_dimensions["A"].width = 24
    for c in range(2, 8):
        ws.column_dimensions[get_column_letter(c)].width = 13

    used = {"Summary"}
    for batch_id, matrix in matrices.items():
        name = batch_names.get(batch_id, "Batch")
        sheet = wb.create_sheet(_safe_sheet_title(name, used))
        _write_matrix_sheet(
            sheet,
            title_lines=[name, f"Register · {_period(start, end)}"],
            matrix=matrix,
        )
    return _xlsx_bytes(wb)


def daily_ledger_xlsx(
    *, brand: str, start: date, end: date, ledger: list[dict], tz_name: str,
) -> bytes:
    """The immutable all-students daily ledger — one row per student per day with
    a record, batch-independent (see daily_service.daily_ledger)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily ledger"

    ws["A1"] = f"{brand} — daily attendance ledger"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = _period(start, end)
    students = len({r["student_id"] for r in ledger})
    ws["A3"] = f"{students} students · {len(ledger)} day records"

    head = 5
    headers = ["#", "Student", "PRN", "Date", "In", "Out", "Status", "Sign-off"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=head, column=c, value=h)
        cell.font = _BOLD
        cell.fill = _HEAD_FILL

    for i, r in enumerate(ledger, start=1):
        row = head + i
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=r["name"])
        ws.cell(row=row, column=3, value=r["enrollment_number"] or "")
        ws.cell(row=row, column=4, value=r["attendance_date"].isoformat())
        ws.cell(row=row, column=5, value=_fmt_time(r["first_in"], tz_name))
        ws.cell(row=row, column=6, value=_fmt_time(r["last_out"], tz_name))
        ws.cell(row=row, column=7, value=r["day_status"])
        ws.cell(row=row, column=8, value=r["signoff"])

    ws.column_dimensions["B"].width = 24
    for col in ("C", "D", "E", "F", "G", "H"):
        ws.column_dimensions[col].width = 12
    return _xlsx_bytes(wb)


# ── PDF (HTML -> headless Chromium) ─────────────────────────────────────────

_PDF_CSS = """
@page { margin: 12mm 10mm; }
* { box-sizing: border-box; }
body { font-family: -apple-system, "Segoe UI", Arial, sans-serif; color:#111; font-size: 10pt; margin:0; }
h1 { font-size: 15pt; margin: 0; }
.sub { color:#555; font-size: 9pt; margin: 2pt 0 10pt; }
table { border-collapse: collapse; width: 100%; }
th, td { border: 1px solid #ccc; padding: 3pt 5pt; font-size: 8.5pt; text-align: left; }
th { background: #eef1f5; }
td.c { text-align: center; }
.P { background:#d7f0db; } .L { background:#fcefc7; } .A { background:#f8d7da; }
.tot td { font-weight: 700; background:#f4f6f9; }
"""


def _doc(body: str) -> str:
    return (
        f"<!DOCTYPE html><html><head><meta charset='utf-8'>"
        f"<style>{_PDF_CSS}</style></head><body>{body}</body></html>"
    )


def _esc(v: Any) -> str:
    return html.escape(str(v))


def _status_class(day_status: str | None) -> str:
    """PDF cell tint class for a day status: P present / L late / A absent."""
    if day_status == "PRESENT":
        return "c P"
    if day_status == "LATE":
        return "c L"
    return "c A"


def student_html(
    *, brand: str, student_name: str, start: date, end: date,
    summary: dict, timeline: list[Any], tz_name: str,
) -> str:
    rows = sorted(timeline, key=lambda r: r.attendance_date)
    body = [
        f"<h1>{_esc(brand)}</h1>",
        f"<p class='sub'>Attendance — {_esc(student_name)} · {_period(start, end)}<br>"
        f"Working days {summary['working_days']} · Present {summary['present_days']} · "
        f"Absent {summary['absent_days']} · <b>{summary['attendance_pct']}%</b></p>",
        "<table><tr><th>Date</th><th>In</th><th>Out</th><th>Status</th><th>Sign-off</th></tr>",
    ]
    for r in rows:
        body.append(
            f"<tr><td>{_esc(r.attendance_date.isoformat())}</td>"
            f"<td>{_esc(_fmt_time(r.first_in, tz_name))}</td>"
            f"<td>{_esc(_fmt_time(r.last_out, tz_name))}</td>"
            f"<td>{_esc(r.day_status)}</td><td>{_esc(r.signoff)}</td></tr>"
        )
    body.append("</table>")
    return _doc("".join(body))


def _matrix_table_html(matrix: dict) -> str:
    dates = matrix["dates"]
    head = "".join(f"<th class='c'>{d.day}/{d.month}</th>" for d in dates)
    parts = [
        f"<table><tr><th>#</th><th>PRN</th><th>Student</th>{head}"
        f"<th>P</th><th>Ab</th><th>%</th></tr>"
    ]
    for i, s in enumerate(matrix["students"], start=1):
        cells = "".join(
            f"<td class='c {c}'>{c}</td>" for c in s["cells"]
        )
        absent = s["working_days"] - s["present"]
        parts.append(
            f"<tr><td>{i}</td><td>{_esc(s['enrollment_number'] or '')}</td>"
            f"<td>{_esc(s['name'])}</td>{cells}"
            f"<td class='c'>{s['present']}</td><td class='c'>{absent}</td>"
            f"<td class='c'>{s['attendance_pct']}</td></tr>"
        )
    tot = "".join(f"<td class='c'>{n}</td>" for n in matrix["day_present"])
    parts.append(
        f"<tr class='tot'><td></td><td></td><td>Present / day</td>{tot}"
        f"<td></td><td></td><td></td></tr>"
    )
    parts.append("</table>")
    return "".join(parts)


def batch_html(
    *, brand: str, batch_name: str, start: date, end: date, matrix: dict,
) -> str:
    body = (
        f"<h1>{_esc(brand)} — {_esc(batch_name)}</h1>"
        f"<p class='sub'>Attendance register · {_period(start, end)} · "
        f"{matrix['student_count']} students · {len(matrix['dates'])} working days "
        f"(P present / L late / A absent)</p>"
        f"{_matrix_table_html(matrix)}"
    )
    return _doc(body)


def all_batches_html(
    *, brand: str, start: date, end: date, summaries: list[dict],
) -> str:
    rows = "".join(
        f"<tr><td>{_esc(s['batch_name'])}</td><td>{_esc(s['batch_code'])}</td>"
        f"<td class='c'>{s['student_count']}</td><td class='c'>{s['working_days']}</td>"
        f"<td class='c'>{s['present']}</td><td class='c'>{s['total_slots']}</td>"
        f"<td class='c'>{s['avg_pct']}</td></tr>"
        for s in summaries
    )
    body = (
        f"<h1>{_esc(brand)} — all batches</h1>"
        f"<p class='sub'>Attendance summary · {_period(start, end)}</p>"
        f"<table><tr><th>Batch</th><th>Code</th><th>Students</th><th>Working days</th>"
        f"<th>Present</th><th>Total</th><th>Avg %</th></tr>{rows}</table>"
    )
    return _doc(body)


def daily_ledger_html(
    *, brand: str, start: date, end: date, ledger: list[dict], tz_name: str,
) -> str:
    students = len({r["student_id"] for r in ledger})
    rows = "".join(
        f"<tr><td class='c'>{i}</td><td>{_esc(r['name'])}</td>"
        f"<td>{_esc(r['enrollment_number'] or '')}</td>"
        f"<td>{_esc(r['attendance_date'].isoformat())}</td>"
        f"<td class='c'>{_esc(_fmt_time(r['first_in'], tz_name))}</td>"
        f"<td class='c'>{_esc(_fmt_time(r['last_out'], tz_name))}</td>"
        f"<td class='{_status_class(r['day_status'])}'>{_esc(r['day_status'])}</td>"
        f"<td>{_esc(r['signoff'])}</td></tr>"
        for i, r in enumerate(ledger, start=1)
    )
    body = (
        f"<h1>{_esc(brand)} — daily attendance ledger</h1>"
        f"<p class='sub'>{_period(start, end)} · {students} students · "
        f"{len(ledger)} day records</p>"
        f"<table><tr><th>#</th><th>Student</th><th>PRN</th><th>Date</th>"
        f"<th>In</th><th>Out</th><th>Status</th><th>Sign-off</th></tr>{rows}</table>"
    )
    return _doc(body)


# ── Day report (single day, single batch — matches the shared sample PDF) ────

# Academy brand mark (mirror of frontend/public/logo.svg) inlined so the headless
# Chromium render is self-contained (no file server for the PDF).
_MSA_LOGO_SVG = (
    "<svg width='40' height='40' viewBox='0 0 32 32' fill='none' "
    "xmlns='http://www.w3.org/2000/svg'>"
    "<rect width='32' height='32' rx='7' fill='#003464'/>"
    "<g stroke='#ffffff' stroke-width='1.6' fill='none'>"
    "<ellipse cx='16' cy='16' rx='9' ry='3.6'/>"
    "<ellipse cx='16' cy='16' rx='9' ry='3.6' transform='rotate(60 16 16)'/>"
    "<ellipse cx='16' cy='16' rx='9' ry='3.6' transform='rotate(120 16 16)'/>"
    "</g><circle cx='16' cy='16' r='2.4' fill='#f4a300'/></svg>"
)

_VENDOR = "EduPulse Technologies"

_DAY_PDF_CSS = """
@page { margin: 12mm 10mm; }
* { box-sizing: border-box; }
body { font-family: -apple-system, "Segoe UI", Arial, sans-serif; color:#111; font-size:10pt; margin:0; }
.brandbar { display:flex; align-items:center; gap:12px; border-bottom:2px solid #003464; padding-bottom:10px; }
.brandbar h1 { font-size:20pt; margin:0; color:#003464; font-weight:800; }
.meta { display:flex; justify-content:space-between; background:#f4f6f9; border:1px solid #dde3ea; border-left:3px solid #003464; padding:7pt 10pt; margin:10px 0; font-size:9.5pt; }
.tiles { display:flex; gap:8px; margin-bottom:10px; }
.tile { flex:1; text-align:center; border:1px solid #e3e8ee; border-radius:6px; padding:8pt; }
.tile .n { font-size:20pt; font-weight:800; }
.tile.total .n { color:#003464; } .tile.present .n { color:#137a52; } .tile.absent .n { color:#c0392b; }
.tile .l { font-size:8pt; color:#666; text-transform:uppercase; letter-spacing:.04em; }
table { border-collapse:collapse; width:100%; }
th, td { border:1px solid #d7dde4; padding:4pt 6pt; font-size:9pt; text-align:left; }
th { background:#0f2947; color:#fff; font-weight:700; }
tr:nth-child(even) td { background:#f7f9fb; }
td.c { text-align:center; }
.st-present { color:#137a52; font-weight:700; }
.st-absent { color:#c0392b; font-weight:700; }
.st-late { color:#b9770e; font-weight:700; }
.manual { display:inline-block; margin-left:4px; font-size:7pt; color:#003464; border:1px solid #003464; border-radius:3px; padding:0 3px; vertical-align:middle; }
.footer { margin-top:12px; background:#f4f6f9; text-align:center; padding:7pt; font-size:9.5pt; border-radius:4px; }
.credit { text-align:right; color:#888; font-size:8pt; font-style:italic; margin-top:6px; }
"""


def _fmt_day_long(d: date) -> str:
    return d.strftime("%d %B %Y").lstrip("0")


def _day_counts(rows: list[dict]) -> tuple[int, int, int, float]:
    total = len(rows)
    present = sum(1 for r in rows if r["mark"] == "P")
    absent = total - present
    pct = round(present / total * 100, 1) if total else 0.0
    return total, present, absent, pct


def _status_cell(row: dict) -> str:
    ds = row.get("day_status")
    manual = "<span class='manual'>Manual</span>" if row.get("source") == "MANUAL" else ""
    if ds == "PRESENT" or ds == "LATE":
        cls = "st-late" if ds == "LATE" else "st-present"
        label = "Late" if ds == "LATE" else "Present"
        return f"<span class='{cls}'>{label}</span>{manual}"
    return f"<span class='st-absent'>Absent</span>{manual}"


def day_report_html(
    *, brand: str, batch_name: str, day: date, rows: list[dict], tz_name: str,
) -> str:
    total, present, absent, pct = _day_counts(rows)
    body = [
        f"<div class='brandbar'>{_MSA_LOGO_SVG}<h1>{_esc(brand)}</h1></div>",
        f"<div class='meta'><span><b>Batch:</b> {_esc(batch_name)}</span>"
        f"<span><b>Date:</b> {_esc(_fmt_day_long(day))}</span>"
        f"<span><b>Generated:</b> {_VENDOR}</span></div>",
        "<div class='tiles'>"
        f"<div class='tile total'><div class='n'>{total}</div><div class='l'>Total Students</div></div>"
        f"<div class='tile present'><div class='n'>{present}</div><div class='l'>Present</div></div>"
        f"<div class='tile absent'><div class='n'>{absent}</div><div class='l'>Absent</div></div>"
        "</div>",
        "<table><tr><th>Sr. No.</th><th>PRN</th><th>Student Name</th><th>RFID</th>"
        "<th>In Time</th><th>Out Time</th><th>Status</th></tr>",
    ]
    for i, r in enumerate(rows, start=1):
        body.append(
            f"<tr><td class='c'>{i}</td>"
            f"<td>{_esc(r.get('enrollment_number') or '—')}</td>"
            f"<td>{_esc(r['name'])}</td>"
            f"<td>{_esc(r.get('rfid_number') or '—')}</td>"
            f"<td class='c'>{_esc(_fmt_time(r.get('first_in'), tz_name))}</td>"
            f"<td class='c'>{_esc(_fmt_time(r.get('last_out'), tz_name))}</td>"
            f"<td>{_status_cell(r)}</td></tr>"
        )
    body.append("</table>")
    body.append(
        f"<div class='footer'>Total Present: <b>{present}</b> | "
        f"Total Absent: <b>{absent}</b> | Attendance: <b>{pct}%</b></div>"
    )
    body.append(f"<div class='credit'>Powered by {_VENDOR}</div>")
    return (
        f"<!DOCTYPE html><html><head><meta charset='utf-8'>"
        f"<style>{_DAY_PDF_CSS}</style></head><body>{''.join(body)}</body></html>"
    )


def day_report_xlsx(
    *, brand: str, batch_name: str, day: date, rows: list[dict], tz_name: str,
) -> bytes:
    total, present, absent, pct = _day_counts(rows)
    wb = Workbook()
    ws = wb.active
    ws.title = "Day report"

    ws["A1"] = brand
    ws["A1"].font = Font(bold=True, size=14, color="003464")
    ws["A2"] = f"Batch: {batch_name}  ·  Date: {_fmt_day_long(day)}  ·  Generated: {_VENDOR}"
    ws["A3"] = f"Total {total}  ·  Present {present}  ·  Absent {absent}  ·  {pct}%"

    head = 5
    headers = ["Sr. No.", "PRN", "Student Name", "RFID", "In Time", "Out Time", "Status"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=head, column=c, value=h)
        cell.font = _BOLD
        cell.fill = _HEAD_FILL

    for i, r in enumerate(rows, start=1):
        row = head + i
        status = "Absent"
        if r.get("day_status") in ("PRESENT", "LATE"):
            status = "Late" if r["day_status"] == "LATE" else "Present"
        if r.get("source") == "MANUAL":
            status += " (Manual)"
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=r.get("enrollment_number") or "")
        ws.cell(row=row, column=3, value=r["name"])
        ws.cell(row=row, column=4, value=r.get("rfid_number") or "")
        ws.cell(row=row, column=5, value=_fmt_time(r.get("first_in"), tz_name))
        ws.cell(row=row, column=6, value=_fmt_time(r.get("last_out"), tz_name))
        ws.cell(row=row, column=7, value=status)

    foot = head + len(rows) + 2
    ws.cell(row=foot, column=1,
            value=f"Total Present: {present} | Total Absent: {absent} | Attendance: {pct}%").font = _BOLD
    ws.cell(row=foot + 1, column=1, value=f"Powered by {_VENDOR}")

    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 24
    for col in ("D", "E", "F", "G"):
        ws.column_dimensions[col].width = 12
    return _xlsx_bytes(wb)


async def render_html_to_pdf(doc_html: str, *, landscape: bool = False) -> bytes:
    """Headless Chromium HTML -> PDF (same engine as the paper PDFs)."""
    from playwright.async_api import async_playwright

    async with async_playwright() as p:
        browser = await p.chromium.launch(args=["--no-sandbox"])
        try:
            page = await browser.new_page()
            await page.set_content(doc_html, wait_until="networkidle")
            return await page.pdf(
                format="A4",
                landscape=landscape,
                print_background=True,
                margin={"top": "12mm", "right": "10mm", "bottom": "12mm", "left": "10mm"},
            )
        finally:
            await browser.close()
