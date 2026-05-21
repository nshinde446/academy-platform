import csv
import io
import uuid
from typing import Any

from fastapi import HTTPException, UploadFile, status
from sqlalchemy.ext.asyncio import AsyncSession

from app.modules.academic.repositories import academic_repository
from app.modules.audit.services import audit_service
from app.modules.teacher.repositories import teacher_repository

# Excel/CSV column -> Teacher field. employee_id / gender are not (yet) on the
# Teacher model — they are accepted and silently ignored so existing
# CSV exports don't break.
COLUMN_MAPPING = {
    "name": "name",
    "first name": "first_name",
    "first_name": "first_name",
    "last name": "last_name",
    "last_name": "last_name",
    "email": "email",
    "phone": "phone",
    "qualification": "qualification",
    "subjects": "subjects",
    "subject": "subjects",
}


def _split_name(value: str) -> tuple[str, str]:
    parts = value.strip().split(maxsplit=1)
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], parts[1]


def _normalize(header: str) -> str:
    return header.strip().lower()


def _parse_csv(content: bytes) -> list[dict[str, Any]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


def _parse_xlsx(content: bytes) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb.active
    if ws is None:
        return []
    rows = ws.iter_rows(values_only=True)
    try:
        headers = [str(h) if h is not None else "" for h in next(rows)]
    except StopIteration:
        return []
    out: list[dict[str, Any]] = []
    for row in rows:
        record = {
            headers[i]: ("" if row[i] is None else str(row[i]))
            for i in range(min(len(headers), len(row)))
        }
        if any(v for v in record.values()):
            out.append(record)
    return out


def _row_to_teacher_kwargs(row: dict[str, Any]) -> tuple[dict[str, Any] | None, list[str]]:
    """Return (teacher kwargs, subject_names). Returns (None, []) if row is unusable."""
    mapped: dict[str, Any] = {}
    subjects_raw: str = ""
    for raw_key, raw_value in row.items():
        if raw_key is None:
            continue
        field = COLUMN_MAPPING.get(_normalize(raw_key))
        if field is None:
            continue
        value = (raw_value or "").strip() if isinstance(raw_value, str) else raw_value
        if value in ("", None):
            continue
        if field == "subjects":
            subjects_raw = str(value)
        else:
            mapped[field] = value

    name_value = mapped.pop("name", None)
    if name_value:
        first, last = _split_name(str(name_value))
        mapped.setdefault("first_name", first)
        mapped.setdefault("last_name", last)

    if not mapped.get("first_name"):
        return None, []
    mapped.setdefault("last_name", "")

    subject_names = [s.strip() for s in subjects_raw.split(",") if s.strip()]
    return mapped, subject_names


async def import_teachers(
    session: AsyncSession,
    file: UploadFile,
    branch_id: uuid.UUID,
    current_user_id: uuid.UUID,
    ip_address: str | None = None,
) -> dict[str, Any]:
    content = await file.read()
    filename = (file.filename or "").lower()

    if filename.endswith(".csv"):
        rows = _parse_csv(content)
    elif filename.endswith((".xlsx", ".xls")):
        rows = _parse_xlsx(content)
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported file type. Use .csv or .xlsx",
        )

    imported = 0
    skipped = 0
    errors: list[str] = []

    for idx, row in enumerate(rows, start=2):  # row 1 is header
        kwargs, subject_names = _row_to_teacher_kwargs(row)
        if kwargs is None:
            skipped += 1
            errors.append(f"Row {idx}: missing required 'Name'")
            continue
        try:
            kwargs["branch_id"] = branch_id
            teacher = await teacher_repository.create(session, **kwargs)
            if subject_names:
                subjects = await academic_repository.find_subjects_by_names(
                    session, branch_id, subject_names
                )
                if subjects:
                    await teacher_repository.add_subject_mappings(
                        session,
                        teacher.id,
                        branch_id,
                        [s.id for s in subjects],
                    )
                found_lower = {s.name.lower() for s in subjects}
                missing = sorted(
                    {n for n in subject_names if n.lower() not in found_lower}
                )
                if missing:
                    errors.append(
                        f"Row {idx}: subjects not found in this branch: {', '.join(missing)}"
                    )
            imported += 1
        except Exception as exc:  # noqa: BLE001
            skipped += 1
            errors.append(f"Row {idx}: {exc}")

    await audit_service.log_action(
        session,
        user_id=current_user_id,
        action="IMPORT",
        table_name="teachers",
        record_id=uuid.uuid4(),
        new_values={"imported": imported, "skipped": skipped},
        ip_address=ip_address,
        branch_id=branch_id,
    )

    return {"imported": imported, "skipped": skipped, "errors": errors}
